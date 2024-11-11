from django.shortcuts import render, redirect
from django.http import HttpResponse
from .models import Usuarios, Divisiones, Procedimientos
from django.contrib import messages
from .forms import *
from .models import *
from django.shortcuts import get_object_or_404
from django.http import JsonResponse
import json
from datetime import datetime
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from itertools import chain
from django.views.decorators.cache import never_cache
import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.db.models import Count
import json
from datetime import datetime
from dateutil.relativedelta import relativedelta
from datetime import timedelta
import os
from django.conf import settings
import subprocess
from django.db.models import Case, When

# Vista Personalizada para el error 404
def custom_404_view(request, exception):
    return render(request, "404.html", status=404)

# Vista para descargar la base de datos
def descargar_base_datos(request):
    # Crear un nombre de archivo basado en la fecha actual
    fecha_actual = timezone.now().strftime('%Y-%m-%d_%H-%M-%S')

    # Comprobamos el tipo de base de datos
    db_engine = settings.DATABASES['default']['ENGINE']
    
    if 'sqlite3' in db_engine:
        # Para SQLite, usamos el archivo .sqlite3 directamente
        db_path = settings.DATABASES['default']['NAME']
        filename = f"backup_{fecha_actual}.sqlite3"
        
        # Abrimos el archivo y lo enviamos como respuesta
        with open(db_path, 'rb') as f:
            response = HttpResponse(f.read(), content_type="application/x-sqlite3")
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
    
    elif 'postgresql' in db_engine:
        # Para PostgreSQL, generamos un archivo .sql con pg_dump
        db_name = settings.DATABASES['default']['NAME']
        db_user = settings.DATABASES['default']['USER']
        db_password = settings.DATABASES['default']['PASSWORD']
        filename = f"backup_{fecha_actual}.sql"

        # Configuramos el comando pg_dump
        dump_cmd = [
            "pg_dump", 
            "-U", db_user, 
            "-d", db_name, 
            "-F", "c"
        ]

        # Ejecutamos el comando y capturamos el resultado
        os.environ['PGPASSWORD'] = db_password
        with subprocess.Popen(dump_cmd, stdout=subprocess.PIPE) as proc:
            response = HttpResponse(proc.stdout, content_type="application/sql")
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
    
    else:
        # Otros motores de base de datos
        return HttpResponse("Motor de base de datos no compatible", status=400)

# Api para crear el excel de exportacion de la tabla
def generar_excel(request):
    # Crear un libro de trabajo y una hoja
    workbook = openpyxl.Workbook()
    hoja = workbook.active
    hoja.title = "Procedimientos"

    # Agregar encabezados a la primera fila
    encabezados = [
        "División", "Solicitante", "Jefe Comisión", "Municipio", 
        "Parroquia", "Fecha", "Hora", "Dirección", 
        "Tipo de Procedimiento", "Detalles", "Persona Presente", "Descripcion"
    ]
    hoja.append(encabezados)

    # Obtener datos de los procedimientos
    procedimientos = Procedimientos.objects.all()

    # Agregar datos a la hoja
    for procedimiento in procedimientos:
        # Determinar los datos de solicitante y jefe de comisión
        solicitante = (f"{procedimiento.id_solicitante.jerarquia} "
                       f"{procedimiento.id_solicitante.nombres} "
                       f"{procedimiento.id_solicitante.apellidos}") \
                      if procedimiento.id_solicitante.id != 0 \
                      else procedimiento.solicitante_externo

        jefe_comision = (f"{procedimiento.id_jefe_comision.jerarquia} "
                         f"{procedimiento.id_jefe_comision.nombres} "
                         f"{procedimiento.id_jefe_comision.apellidos}") \
                        if procedimiento.id_jefe_comision.id != 0 else ""

        # Recopilar datos de "Personas Presentes" desde todas las tablas relacionadas
        personas_presentes = []
        detalles_procedimientos = []
        descripciones_proc = []

        # Abastecimiento_agua
        for detalles in procedimiento.abastecimiento_agua_set.all():
            personas_presentes.append(f"{detalles.nombres} {detalles.apellidos} {detalles.cedula}")
            detalles_procedimientos.append(f"{detalles.id_tipo_servicio.nombre_institucion}")
            descripciones_proc.append(detalles.descripcion)

        # Apoyo_Unidades
        for detalles in procedimiento.apoyo_unidades_set.all():
            detalles_procedimientos.append(detalles.id_tipo_apoyo.tipo_apoyo)
            descripciones_proc.append(detalles.descripcion)

        # Guardia_Prevencion
        for detalles in procedimiento.guardia_prevencion_set.all():
            detalles_procedimientos.append(detalles.id_motivo_prevencion.motivo)
            descripciones_proc.append(detalles.descripcion)

        # Atendido_no_Efectuado
        for detalles in procedimiento.atendido_no_efectuado_set.all():
            descripciones_proc.append(detalles.descripcion)

        # Despliegue_Seguridad
        for detalles in procedimiento.despliegue_seguridad_set.all():
            detalles_procedimientos.append(detalles.motivo_despliegue.motivo)
            descripciones_proc.append(detalles.descripcion)

        # Fallecidos
        for detalles in procedimiento.fallecidos_set.all():
            personas_presentes.append(f"{detalles.nombres} {detalles.apellidos} {detalles.cedula}")
            detalles_procedimientos.append(detalles.motivo_fallecimiento)
            descripciones_proc.append(detalles.descripcion)

        # Falsa_Alarma
        for detalles in procedimiento.falsa_alarma_set.all():
            detalles_procedimientos.append(detalles.motivo_alarma.motivo)
            descripciones_proc.append(detalles.descripcion)

        # Servicios_Especiales
        for detalles in procedimiento.servicios_especiales_set.all():
            detalles_procedimientos.append(detalles.tipo_servicio.serv_especiales)
            descripciones_proc.append(detalles.descripcion)

        for rescate in procedimiento.rescate_set.all():
            # Añadir el tipo de rescate al listado de detalles
            detalles_procedimientos.append(rescate.tipo_rescate.tipo_rescate)
            
            # Detalles de personas rescatadas
            detalles_personas = []
            for persona in rescate.rescate_persona_set.all():
                detalles_personas.append(f"{persona.nombre} {persona.apellidos} {persona.cedula}")
                descripciones_proc.append(persona.descripcion)
            
            # Añadir detalles de personas rescatadas, si existen
            if detalles_personas:
                personas_presentes.append(f"{''.join(detalles_personas)}")

            # Detalles de animales rescatados
            detalles_animales = []
            for animal in rescate.rescate_animal_set.all():
                detalles_animales.append(f"Especie: {animal.especie}")
                descripciones_proc.append(animal.descripcion)
            
            # Añadir detalles de animales rescatados, si existen
            if detalles_animales:
                detalles_procedimientos.append(f"{''.join(detalles_animales)}")

        # Incendios -> Persona_Presente
        for incendio in procedimiento.incendios_set.all():
            detalles_procedimientos.append(incendio.id_tipo_incendio.tipo_incendio)
            descripciones_proc.append(incendio.descripcion)
            for persona in incendio.persona_presente_set.all():
                personas_presentes.append(f"{persona.nombre} {persona.apellidos} {persona.cedula}")

       # Atenciones Paramédicas
        for atencion in procedimiento.atenciones_paramedicas_set.all():
            # Añadir el tipo de atención
            detalles_procedimientos.append(atencion.tipo_atencion)

            # Detalles de Emergencias Médicas
            detalles_emergencias = []
            for emergencia in atencion.emergencias_medicas_set.all():
                detalles_emergencias.append(f"{emergencia.nombres} {emergencia.apellidos} {emergencia.cedula}")
                descripciones_proc.append(emergencia.descripcion)
            
            # Añadir detalles de emergencias médicas, si existen
            if detalles_emergencias:
                personas_presentes.append(f"{''.join(detalles_emergencias)}")

            # Detalles de Accidentes de Tránsito
            for accidente in atencion.accidentes_transito_set.all():
                # Añadir tipo de accidente
                detalles_procedimientos.append(accidente.tipo_de_accidente.tipo_accidente)
                
                # Detalles de Lesionados
                detalles_lesionados = []
                for lesionado in accidente.lesionados_set.all():
                    detalles_lesionados.append(f"{lesionado.nombres} {lesionado.apellidos} {lesionado.cedula}, ")
                    descripciones_proc.append(lesionado.descripcion)
                
                # Añadir detalles de lesionados, si existen
                if detalles_lesionados:
                    personas_presentes.append(f"{''.join(detalles_lesionados)}")

        # Traslado_Prehospitalaria
        for traslado in procedimiento.traslado_prehospitalaria_set.all():
            detalles_procedimientos.append(traslado.id_tipo_traslado.tipo_traslado)
            personas_presentes.append(f"{traslado.nombre} {traslado.apellido} {traslado.cedula}")
            descripciones_proc.append(traslado.descripcion)

        # Evaluacion_Riesgo -> Persona_Presente_Eval
        for evaluacion in procedimiento.evaluacion_riesgo_set.all():
            detalles_procedimientos.append(evaluacion.id_tipo_riesgo.tipo_riesgo)
            detalles_procedimientos.append(evaluacion.tipo_estructura)
            descripciones_proc.append(evaluacion.descripcion)
            for persona in evaluacion.persona_presente_eval_set.all():
                personas_presentes.append(f"{persona.nombre} {persona.apellidos} {persona.cedula}")

        # Mitigacion_Riesgos
        for mitigacion in procedimiento.mitigacion_riesgos_set.all():
            detalles_procedimientos.append(mitigacion.id_tipo_servicio.tipo_servicio)
            descripciones_proc.append(mitigacion.descripcion)

        # Puesto_Avanzada
        for avanzada in procedimiento.puesto_avanzada_set.all():
            detalles_procedimientos.append(avanzada.id_tipo_servicio.tipo_servicio)
            descripciones_proc.append(avanzada.descripcion)

        # Asesoramiento
        for asesoramiento in procedimiento.asesoramiento_set.all():
            detalles_procedimientos.append(f"Comercio: {asesoramiento.nombre_comercio} {asesoramiento.rif_comercio}")
            personas_presentes.append(f"{asesoramiento.nombres} {asesoramiento.apellidos} {asesoramiento.cedula}")
            descripciones_proc.append(asesoramiento.descripcion)

        # Reinspeccion_Prevencion
        for detalles in procedimiento.reinspeccion_prevencion_set.all():
            detalles_procedimientos.append(f"Comercio: {detalles.nombre_comercio} {detalles.rif_comercio}")
            personas_presentes.append(f"{detalles.nombre} {detalles.apellidos} {detalles.cedula}")
            descripciones_proc.append(detalles.descripcion)

        # Artificios_Pirotecnicos
        for artificio in procedimiento.artificios_pirotecnicos_set.all():
            detalles_procedimientos.append(f"{artificio.tipo_procedimiento.tipo}")
            
            # Incendios_Art -> Persona_Presente_Art
            for incendio in artificio.incendios_art_set.all():
                descripciones_proc.append(incendio.descripcion)
                for persona in incendio.persona_presente_art_set.all():
                    personas_presentes.append(f"{persona.nombre} {persona.apellidos} {persona.cedula}")

            # Lesionados_Art
            for lesionado in artificio.lesionados_art_set.all():
                descripciones_proc.append(lesionado.descripcion)
                personas_presentes.append(f"{lesionado.nombres} {lesionado.apellidos} {lesionado.cedula}")

            # Fallecidos_Art
            for fallecido in artificio.fallecidos_art_set.all():
                descripciones_proc.append(fallecido.descripcion)
                personas_presentes.append(f"{fallecido.nombres} {fallecido.apellidos} {fallecido.cedula}")

        # Inspeccion_Establecimiento_Art
        for inspeccion in procedimiento.inspeccion_establecimiento_art_set.all():
            detalles_procedimientos.append(f"Comercio: {inspeccion.nombre_comercio} {inspeccion.rif_comercio}")
            descripciones_proc.append(inspeccion.descripcion)
            personas_presentes.append(f"{inspeccion.encargado_nombre} {inspeccion.encargado_apellidos} {inspeccion.encargado_cedula}")

        # Valoracion_Medica
        for valoracion in procedimiento.valoracion_medica_set.all():
            detalles_procedimientos.append(procedimiento.tipo_servicio)
            personas_presentes.append(f"{valoracion.nombre} {valoracion.apellido} {valoracion.cedula}")
            descripciones_proc.append(valoracion.descripcion)

        # Detalles_Enfermeria
        for enfermeria in procedimiento.detalles_enfermeria_set.all():
            detalles_procedimientos.append(procedimiento.dependencia)
            personas_presentes.append(f"{enfermeria.nombre} {enfermeria.apellido} {enfermeria.cedula}")
            descripciones_proc.append(enfermeria.descripcion)

        # Procedimientos_Psicologia
        for psicologia in procedimiento.procedimientos_psicologia_set.all():
            detalles_procedimientos.append("Consultas Psicologicas")
            personas_presentes.append(f"{psicologia.nombre} {psicologia.apellido} {psicologia.cedula}")
            descripciones_proc.append(psicologia.descripcion)

        # Procedimientos_Capacitacion
        for capacitacion in procedimiento.procedimientos_capacitacion_set.all():
            detalles_procedimientos.append(f"Capacitación: {capacitacion.tipo_capacitacion} - {capacitacion.descripcion}")
            descripciones_proc.append(capacitacion.descripcion)

        # Procedimientos_Frente_Preventivo
        for frente_preventivo in procedimiento.procedimientos_frente_preventivo_set.all():
            detalles_procedimientos.append(f"Actividad: {frente_preventivo.nombre_actividad} - {frente_preventivo.descripcion}")
            descripciones_proc.append(frente_preventivo.descripcion)

        # Datos de la Jornada Médica
        for jornada in procedimiento.jornada_medica_set.all():
            detalles_procedimientos.append(f"{jornada.nombre_jornada} - {jornada.cant_personas_aten}")
            descripciones_proc.append(f"{jornada.descripcion}")  # Incluyendo la cantidad

        # Inspecciones
        for inspeccion in procedimiento.inspeccion_prevencion_asesorias_tecnicas_set.all():
            detalles_procedimientos.append(inspeccion.tipo_inspeccion)
            personas_presentes.append(f"{inspeccion.persona_sitio_nombre} {inspeccion.persona_sitio_apellido} ({inspeccion.persona_sitio_cedula})")
            descripciones_proc.append(inspeccion.descripcion)

        for inspeccion in procedimiento.inspeccion_habitabilidad_set.all():
            detalles_procedimientos.append(inspeccion.tipo_inspeccion)
            personas_presentes.append(f"{inspeccion.persona_sitio_nombre} {inspeccion.persona_sitio_apellido} ({inspeccion.persona_sitio_cedula})")
            descripciones_proc.append(inspeccion.descripcion)

        for inspeccion in procedimiento.inspeccion_otros_set.all():
            detalles_procedimientos.append(inspeccion.tipo_inspeccion)
            personas_presentes.append(f"{inspeccion.persona_sitio_nombre} {inspeccion.persona_sitio_apellido} ({inspeccion.persona_sitio_cedula})")
            descripciones_proc.append(inspeccion.descripcion)

        for inspeccion in procedimiento.inspeccion_arbol_set.all():
            detalles_procedimientos.append(f"{inspeccion.tipo_inspeccion} -- {inspeccion.especie} ({inspeccion.altura_aprox})")
            personas_presentes.append(f"{inspeccion.persona_sitio_nombre} {inspeccion.persona_sitio_apellido} {inspeccion.persona_sitio_cedula}")
            descripciones_proc.append(inspeccion.descripcion)

        # Para cada registro de 'Investigacion' asociado con 'Procedimientos'
        for investigacion in procedimiento.investigacion_set.all():
            # Añadir el tipo de investigación y el tipo de siniestro
            detalles_procedimientos.append(f"{investigacion.id_tipo_investigacion.tipo_investigacion} - {investigacion.tipo_siniestro}")
            
            # Investigacion -> Investigacion_Vehiculo
            detalles_propietario = []
            for vehiculo in investigacion.investigacion_vehiculo_set.all():
                detalles_propietario.append(f"{vehiculo.nombre_propietario} {vehiculo.apellido_propietario} {vehiculo.cedula_propietario}")
                descripciones_proc.append(vehiculo.descripcion)
            if detalles_propietario:
                personas_presentes.append(f"{''.join(detalles_propietario)}")
            
            # Investigacion -> Investigacion_Comercio
            detalles_comercios = []
            for comercio in investigacion.investigacion_comercio_set.all():
                detalles_comercios.append(f"{comercio.nombre_propietario} {comercio.apellido_propietario} {comercio.cedula_propietario}")
                descripciones_proc.append(comercio.descripcion)
            if detalles_comercios:
                personas_presentes.append(f"{''.join(detalles_comercios)}")
            
            # Investigacion -> Investigacion_Estructura_Vivienda
            detalles_estructuras = []
            for estructura in investigacion.investigacion_estructura_vivienda_set.all():
                detalles_estructuras.append(f"{estructura.nombre} {estructura.apellido} {estructura.cedula}")
                descripciones_proc.append(estructura.descripcion)
            if detalles_estructuras:
                personas_presentes.append(f"{''.join(detalles_estructuras)}")

        # Convertir lista de personas presentes a string separado por punto y coma
        personas_presentes_str = " -- ".join(personas_presentes)
        detalles_str = " -- ".join(detalles_procedimientos)  # Convertir descripciones a string
        descripcion_str = " -- ".join(descripciones_proc)  # Convertir descripciones a string

        # Agregar la fila de datos
        hoja.append([
            procedimiento.id_division.division,
            solicitante,
            jefe_comision,
            procedimiento.id_municipio.municipio,
            procedimiento.id_parroquia.parroquia,
            procedimiento.fecha,
            procedimiento.hora,
            procedimiento.direccion,
            procedimiento.id_tipo_procedimiento.tipo_procedimiento,
            detalles_str,
            personas_presentes_str,
            descripcion_str,
        ])

    # Ajustar el ancho de las columnas
    for column in hoja.columns:
        max_length = max(len(str(cell.value)) for cell in column if cell.value) + 2
        hoja.column_dimensions[get_column_letter(column[0].column)].width = max_length

    # Configurar la respuesta HTTP para descargar el archivo
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = "attachment; filename=procedimientos.xlsx"
    workbook.save(response)
    return response

# Api para crear el excel de exportacion de la tabla
def generar_excel_personal(request):
    # Crear un libro de trabajo y una hoja
    workbook = openpyxl.Workbook()
    hoja = workbook.active
    hoja.title = "Procedimientos"

    # Agregar encabezados a la primera fila
    encabezados = [
        "Nombres", "Apellidos", "Jerarquia", "Cargo", 
        "Cedula", "Sexo", "Contrato", "Estado"
    ]
    hoja.append(encabezados)

    # Obtener datos de los procedimientos
    procedimientos = Personal.objects.exclude(id__in=[0, 4])

    # Agregar datos a la hoja
    for procedimiento in procedimientos:
        # Agregar la fila de datos
        hoja.append([
            procedimiento.nombres,
            procedimiento.apellidos,
            procedimiento.jerarquia,
            procedimiento.cargo,
            procedimiento.cedula,
            procedimiento.sexo,
            procedimiento.rol,
            procedimiento.status,
        ])

    # Ajustar el ancho de las columnas
    for column in hoja.columns:
        max_length = max(len(str(cell.value)) for cell in column if cell.value) + 2
        hoja.column_dimensions[get_column_letter(column[0].column)].width = max_length

    # Configurar la respuesta HTTP para descargar el archivo
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = "attachment; filename=procedimientos.xlsx"
    workbook.save(response)
    return response

# Api para crear seccion de lista de procedimientos por cada division, por tipo y parroquia en la seccion de Estadistica
def generar_resultados(request):
    try:
        month = request.GET.get("month")
    
        # Filtrar procedimientos del año actual si no se especifica mes
        if month:
            year, month = map(int, month.split("-"))
            procedimientos = Procedimientos.objects.filter(fecha__year=year, fecha__month=month)
        else:
            current_year = datetime.now().year
            procedimientos = Procedimientos.objects.filter(fecha__year=current_year)

        # Diccionario para almacenar los resultados por división
        resultados_divisiones = {}

        # Definir los nombres de las divisiones
        nombres_divisiones = {
            1: 'Rescate',
            2: 'Operaciones',
            3: 'Prevención',
            4: 'Grumae',
            5: 'Prehospitalaria',
            6: 'Enfermería',
            7: 'ServiciosMédicos',
            8: 'Psicología',
            9: 'Capacitación'
        }

        for division_id, division_nombre in nombres_divisiones.items():
            # Filtrar procedimientos por división
            division_procedimientos = procedimientos.filter(id_division=division_id)

            # Agrupar y contar procedimientos por tipo y parroquia
            tipos_procedimientos_parroquias = (
                division_procedimientos
                .values('id_tipo_procedimiento__tipo_procedimiento', 'id_parroquia__parroquia')
                .annotate(cantidad=Count('id'))
                .order_by('id_tipo_procedimiento__tipo_procedimiento', 'id_parroquia__parroquia')
            )

            # Estructura de resultados por división, tipo de procedimiento y parroquia
            resultados_divisiones[division_nombre] = {
                'total_por_tipo': {},  # Añadir un diccionario para los totales por tipo
                'detalles': {}  # Añadir un diccionario para los detalles
            }

            for item in tipos_procedimientos_parroquias:
                tipo_procedimiento = item['id_tipo_procedimiento__tipo_procedimiento']
                parroquia = item['id_parroquia__parroquia']
                cantidad = item['cantidad']

                # Añadir al diccionario de detalles
                if tipo_procedimiento not in resultados_divisiones[division_nombre]['detalles']:
                    resultados_divisiones[division_nombre]['detalles'][tipo_procedimiento] = {}

                resultados_divisiones[division_nombre]['detalles'][tipo_procedimiento][parroquia] = cantidad

                # Sumar al total por tipo
                if tipo_procedimiento not in resultados_divisiones[division_nombre]['total_por_tipo']:
                    resultados_divisiones[division_nombre]['total_por_tipo'][tipo_procedimiento] = 0

                resultados_divisiones[division_nombre]['total_por_tipo'][tipo_procedimiento] += cantidad

        # Convertir a JSON y devolver como respuesta
        return JsonResponse(resultados_divisiones, json_dumps_params={'ensure_ascii': False, 'indent': 4})
    
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

# Api para crear seccion de grafica anual de meses en el dashboard
def filtrado_mes(mes):
    año_actual = datetime.now().year

    # Filtrado de datos por mes para la grafica anual
    procedimientos_mes = Procedimientos.objects.filter(
    fecha__month=mes,  # Septiembre
    fecha__year=año_actual  # Año 2023
    )
    return procedimientos_mes.count()

def obtener_meses(request):
    enero = filtrado_mes(1)
    febrero = filtrado_mes(2)
    marzo = filtrado_mes(3)
    abril = filtrado_mes(4)
    mayo = filtrado_mes(5)
    junio = filtrado_mes(6)
    julio = filtrado_mes(7)
    agosto = filtrado_mes(8)
    septiembre = filtrado_mes(9)
    octubre = filtrado_mes(10)
    noviembre = filtrado_mes(11)
    diciembre = filtrado_mes(12)

    data = {
        "enero": enero,
        "febrero": febrero,
        "marzo": marzo,
        "abril": abril,
        "mayo": mayo,
        "junio": junio,
        "julio": julio,
        "agosto": agosto,
        "septiembre": septiembre,
        "octubre": octubre,
        "noviembre": noviembre,
        "diciembre": diciembre,
    }
    return JsonResponse(data)

# Api para obtener porcentajes para las cards de la seccion del dashboard
def obtener_porcentajes(request, periodo="general"):
    if periodo == "mes":
        now = timezone.now()
        start_of_month = now.replace(day=1)
        procedimientos_queryset = Procedimientos.objects.filter(fecha__gte=start_of_month)
    else:
        procedimientos_queryset = Procedimientos.objects.all()

    # Contar procedimientos por división
    divisiones = {
        'rescate': procedimientos_queryset.filter(id_division=1).count(),
        'operaciones': procedimientos_queryset.filter(id_division=2).count(),
        'prevencion': procedimientos_queryset.filter(id_division=3).count(),
        'grumae': procedimientos_queryset.filter(id_division=4).count(),
        'prehospitalaria': procedimientos_queryset.filter(id_division=5).count(),
        'enfermeria': procedimientos_queryset.filter(id_division=6).count(),
        'servicios_medicos': procedimientos_queryset.filter(id_division=7).count(),
        'psicologia': procedimientos_queryset.filter(id_division=8).count(),
        'capacitacion': procedimientos_queryset.filter(id_division=9).count(),
    }

    # Total de procedimientos
    procedimientos_totales = sum(divisiones.values())


    # Inicializar porcentajes
    porcentajes = {key: 0.0 for key in divisiones.keys()}

    # Calcular y ajustar los porcentajes
    if procedimientos_totales > 0:
        total_ajustado = 0.0

        for key, count in divisiones.items():
            porcentaje = (count / procedimientos_totales) * 100
            porcentaje_redondeado = round(porcentaje, 1)
            porcentajes[key] = porcentaje_redondeado
            total_ajustado += porcentaje_redondeado

        # Calcular el ajuste necesario
        ajuste = 100 - total_ajustado


        if ajuste != 0:
            # Ajustar el último porcentaje
            last_key = list(porcentajes.keys())[-1]
            porcentajes[last_key] += ajuste

            # Asegurarse de que el último porcentaje no exceda 100
            if porcentajes[last_key] > 100:
                porcentajes[last_key] = 100

    return JsonResponse(porcentajes)

# Api para obtener los valores de las cards por parroquias de la seccion del Dashboard
def obtener_procedimientos_parroquias(request):

    username = request.headers.get("X-User-Name")
    
    # Obtener la fecha de hoy y el primer día del mes
    hoy = datetime.now().date()
    primer_dia_mes = hoy.replace(day=1)

    if username == "Sala_Situacional" or username == "Comandancia" or username == "2dacomandancia" or username == "SeRvEr":

        procedimientos = {
            "otros_municipios": {
                "total": Procedimientos.objects.filter(id_parroquia=0).count(),
                "del_mes": Procedimientos.objects.filter(id_parroquia=0, fecha__gte=primer_dia_mes).count(),
                "hoy": Procedimientos.objects.filter(id_parroquia=0, fecha=hoy).count(),
            },
            "concordia": {
                "total": Procedimientos.objects.filter(id_parroquia=1).count(),
                "del_mes": Procedimientos.objects.filter(id_parroquia=1, fecha__gte=primer_dia_mes).count(),
                "hoy": Procedimientos.objects.filter(id_parroquia=1, fecha=hoy).count(),
            },
            "pedro_m": {
                "total": Procedimientos.objects.filter(id_parroquia=2).count(),
                "del_mes": Procedimientos.objects.filter(id_parroquia=2, fecha__gte=primer_dia_mes).count(),
                "hoy": Procedimientos.objects.filter(id_parroquia=2, fecha=hoy).count(),
            },
            "san_juan": {
                "total": Procedimientos.objects.filter(id_parroquia=3).count(),
                "del_mes": Procedimientos.objects.filter(id_parroquia=3, fecha__gte=primer_dia_mes).count(),
                "hoy": Procedimientos.objects.filter(id_parroquia=3, fecha=hoy).count(),
            },
            "san_sebastian": {
                "total": Procedimientos.objects.filter(id_parroquia=4).count(),
                "del_mes": Procedimientos.objects.filter(id_parroquia=4, fecha__gte=primer_dia_mes).count(),
                "hoy": Procedimientos.objects.filter(id_parroquia=4, fecha=hoy).count(),
            },
            "francisco_romero_lobo": {
                "total": Procedimientos.objects.filter(id_parroquia=6).count(),
                "del_mes": Procedimientos.objects.filter(id_parroquia=6, fecha__gte=primer_dia_mes).count(),
                "hoy": Procedimientos.objects.filter(id_parroquia=6, fecha=hoy).count(),
            },
        }
    
    if username == "Operaciones01":
        procedimientos = {
            "otros_municipios": {
                "hoy": Procedimientos.objects.filter(id_parroquia=0, fecha=hoy, id_division=2).count(),
            },
            "concordia": {
                "hoy": Procedimientos.objects.filter(id_parroquia=1, fecha=hoy, id_division=2).count(),
            },
            "pedro_m": {
                "hoy": Procedimientos.objects.filter(id_parroquia=2, fecha=hoy, id_division=2).count(),
            },
            "san_juan": {
                "hoy": Procedimientos.objects.filter(id_parroquia=3, fecha=hoy, id_division=2).count(),
            },
            "san_sebastian": {
                "hoy": Procedimientos.objects.filter(id_parroquia=4, fecha=hoy, id_division=2).count(),
            },
            "francisco_romero_lobo": {
                "hoy": Procedimientos.objects.filter(id_parroquia=6, fecha=hoy, id_division=2).count(),
            },
        }
    
    if username == "Grumae02":
        procedimientos = {
            "otros_municipios": {
                "hoy": Procedimientos.objects.filter(id_parroquia=0, fecha=hoy, id_division=4).count(),
            },
            "concordia": {
                "hoy": Procedimientos.objects.filter(id_parroquia=1, fecha=hoy, id_division=4).count(),
            },
            "pedro_m": {
                "hoy": Procedimientos.objects.filter(id_parroquia=2, fecha=hoy, id_division=4).count(),
            },
            "san_juan": {
                "hoy": Procedimientos.objects.filter(id_parroquia=3, fecha=hoy, id_division=4).count(),
            },
            "san_sebastian": {
                "hoy": Procedimientos.objects.filter(id_parroquia=4, fecha=hoy, id_division=4).count(),
            },
            "francisco_romero_lobo": {
                "hoy": Procedimientos.objects.filter(id_parroquia=6, fecha=hoy, id_division=4).count(),
            },
        }

    if username == "Rescate03":
        procedimientos = {
            "otros_municipios": {
                "hoy": Procedimientos.objects.filter(id_parroquia=0, fecha=hoy, id_division=1).count(),
            },
            "concordia": {
                "hoy": Procedimientos.objects.filter(id_parroquia=1, fecha=hoy, id_division=1).count(),
            },
            "pedro_m": {
                "hoy": Procedimientos.objects.filter(id_parroquia=2, fecha=hoy, id_division=1).count(),
            },
            "san_juan": {
                "hoy": Procedimientos.objects.filter(id_parroquia=3, fecha=hoy, id_division=1).count(),
            },
            "san_sebastian": {
                "hoy": Procedimientos.objects.filter(id_parroquia=4, fecha=hoy, id_division=1).count(),
            },
            "francisco_romero_lobo": {
                "hoy": Procedimientos.objects.filter(id_parroquia=6, fecha=hoy, id_division=1).count(),
            },
        }

    if username == "Prehospitalaria04":
        procedimientos = {
            "otros_municipios": {
                "hoy": Procedimientos.objects.filter(id_parroquia=0, fecha=hoy, id_division=5).count(),
            },
            "concordia": {
                "hoy": Procedimientos.objects.filter(id_parroquia=1, fecha=hoy, id_division=5).count(),
            },
            "pedro_m": {
                "hoy": Procedimientos.objects.filter(id_parroquia=2, fecha=hoy, id_division=5).count(),
            },
            "san_juan": {
                "hoy": Procedimientos.objects.filter(id_parroquia=3, fecha=hoy, id_division=5).count(),
            },
            "san_sebastian": {
                "hoy": Procedimientos.objects.filter(id_parroquia=4, fecha=hoy, id_division=5).count(),
            },
            "francisco_romero_lobo": {
                "hoy": Procedimientos.objects.filter(id_parroquia=6, fecha=hoy, id_division=5).count(),
            },
        }

    if username == "Prevencion05":
        procedimientos = {
            "otros_municipios": {
                "hoy": Procedimientos.objects.filter(id_parroquia=0, fecha=hoy, id_division=3).count(),
            },
            "concordia": {
                "hoy": Procedimientos.objects.filter(id_parroquia=1, fecha=hoy, id_division=3).count(),
            },
            "pedro_m": {
                "hoy": Procedimientos.objects.filter(id_parroquia=2, fecha=hoy, id_division=3).count(),
            },
            "san_juan": {
                "hoy": Procedimientos.objects.filter(id_parroquia=3, fecha=hoy, id_division=3).count(),
            },
            "san_sebastian": {
                "hoy": Procedimientos.objects.filter(id_parroquia=4, fecha=hoy, id_division=3).count(),
            },
            "francisco_romero_lobo": {
                "hoy": Procedimientos.objects.filter(id_parroquia=6, fecha=hoy, id_division=3).count(),
            },
        }

    if username == "Serviciosmedicos06":
        procedimientos = {
            "otros_municipios": {
                "hoy": Procedimientos.objects.filter(id_parroquia=0, fecha=hoy, id_division=7).count(),
            },
            "concordia": {
                "hoy": Procedimientos.objects.filter(id_parroquia=1, fecha=hoy, id_division=7).count(),
            },
            "pedro_m": {
                "hoy": Procedimientos.objects.filter(id_parroquia=2, fecha=hoy, id_division=7).count(),
            },
            "san_juan": {
                "hoy": Procedimientos.objects.filter(id_parroquia=3, fecha=hoy, id_division=7).count(),
            },
            "san_sebastian": {
                "hoy": Procedimientos.objects.filter(id_parroquia=4, fecha=hoy, id_division=7).count(),
            },
            "francisco_romero_lobo": {
                "hoy": Procedimientos.objects.filter(id_parroquia=6, fecha=hoy, id_division=7).count(),
            },
        }

    if username == "Capacitacion07":
        procedimientos = {
            "otros_municipios": {
                "hoy": Procedimientos.objects.filter(id_parroquia=0, fecha=hoy, id_division=9).count(),
            },
            "concordia": {
                "hoy": Procedimientos.objects.filter(id_parroquia=1, fecha=hoy, id_division=9).count(),
            },
            "pedro_m": {
                "hoy": Procedimientos.objects.filter(id_parroquia=2, fecha=hoy, id_division=9).count(),
            },
            "san_juan": {
                "hoy": Procedimientos.objects.filter(id_parroquia=3, fecha=hoy, id_division=9).count(),
            },
            "san_sebastian": {
                "hoy": Procedimientos.objects.filter(id_parroquia=4, fecha=hoy, id_division=9).count(),
            },
            "francisco_romero_lobo": {
                "hoy": Procedimientos.objects.filter(id_parroquia=6, fecha=hoy, id_division=9).count(),
            },
        }

    if username == "Enfermeria08":
        procedimientos = {
            "otros_municipios": {
                "hoy": Procedimientos.objects.filter(id_parroquia=0, fecha=hoy, id_division=6).count(),
            },
            "concordia": {
                "hoy": Procedimientos.objects.filter(id_parroquia=1, fecha=hoy, id_division=6).count(),
            },
            "pedro_m": {
                "hoy": Procedimientos.objects.filter(id_parroquia=2, fecha=hoy, id_division=6).count(),
            },
            "san_juan": {
                "hoy": Procedimientos.objects.filter(id_parroquia=3, fecha=hoy, id_division=6).count(),
            },
            "san_sebastian": {
                "hoy": Procedimientos.objects.filter(id_parroquia=4, fecha=hoy, id_division=6).count(),
            },
            "francisco_romero_lobo": {
                "hoy": Procedimientos.objects.filter(id_parroquia=6, fecha=hoy, id_division=6).count(),
            },
        }

    if username == "Psicologia09":
        procedimientos = {
            "otros_municipios": {
                "hoy": Procedimientos.objects.filter(id_parroquia=0, fecha=hoy, id_division=8).count(),
            },
            "concordia": {
                "hoy": Procedimientos.objects.filter(id_parroquia=1, fecha=hoy, id_division=8).count(),
            },
            "pedro_m": {
                "hoy": Procedimientos.objects.filter(id_parroquia=2, fecha=hoy, id_division=8).count(),
            },
            "san_juan": {
                "hoy": Procedimientos.objects.filter(id_parroquia=3, fecha=hoy, id_division=8).count(),
            },
            "san_sebastian": {
                "hoy": Procedimientos.objects.filter(id_parroquia=4, fecha=hoy, id_division=8).count(),
            },
            "francisco_romero_lobo": {
                "hoy": Procedimientos.objects.filter(id_parroquia=6, fecha=hoy, id_division=8).count(),
            },
        }

    # print(procedimientos)
    return JsonResponse(procedimientos)

# Api para generar valores para las graficas de pie de la seccion de estadistica
def api_procedimientos_division(request):
    division_id = request.GET.get('division_id')
    mes = request.GET.get('mes')

    # Filtrar por división
    procedimientos = Procedimientos.objects.filter(id_division=division_id)

    # Filtrar por mes si se proporciona
    if mes:
        # Convertir 'mes' a un rango de fechas
        fecha_inicio = datetime.strptime(mes, '%Y-%m').date()
        fecha_fin = fecha_inicio.replace(day=1) + relativedelta(months=1)  # Primer día del siguiente mes
        procedimientos = procedimientos.filter(fecha__gte=fecha_inicio, fecha__lt=fecha_fin)

    # Agrupar por tipo de procedimiento
    conteo_procedimientos = procedimientos.values('id_tipo_procedimiento__tipo_procedimiento').annotate(count=Count('id')).order_by('id_tipo_procedimiento__tipo_procedimiento')

    return JsonResponse(list(conteo_procedimientos), safe=False)

# Api para generar valores para las graficas de donut de la seccion de estadistica
def api_procedimientos_division_parroquias(request):
    division_id = request.GET.get('division_id')
    mes = request.GET.get('mes')

    # Filtrar por división
    procedimientos = Procedimientos.objects.filter(id_division=division_id)

    # Filtrar por mes si se proporciona
    if mes:
        # Convertir 'mes' a un rango de fechas
        fecha_inicio = datetime.strptime(mes, '%Y-%m').date()
        fecha_fin = fecha_inicio.replace(day=1) + relativedelta(months=1)  # Primer día del siguiente mes
        procedimientos = procedimientos.filter(fecha__gte=fecha_inicio, fecha__lt=fecha_fin)

    # Agrupar por parroquia y contar
    conteo_procedimientos = procedimientos.values(
        'id_parroquia__parroquia'  # Agrupar por el campo de la parroquia
    ).annotate(count=Count('id')).order_by('id_parroquia__parroquia')

    # Convertir a lista y retornar
    return JsonResponse(list(conteo_procedimientos), safe=False)

# Api para generar valores para la grafica de procedimientos por tipo
def api_procedimientos_tipo(request):
    tipo_procedimiento_id = request.GET.get('tipo_procedimiento_id')
    mes = request.GET.get('mes')

    # Filtrar por tipo de procedimiento
    procedimientos = Procedimientos.objects.all()
    if tipo_procedimiento_id:
        procedimientos = procedimientos.filter(id_tipo_procedimiento=tipo_procedimiento_id)

    # Filtrar por mes si se proporciona
    if mes:
        fecha_inicio = datetime.strptime(mes, '%Y-%m').date()
        fecha_fin = fecha_inicio.replace(day=1) + relativedelta(months=1)
        procedimientos = procedimientos.filter(fecha__gte=fecha_inicio, fecha__lt=fecha_fin)

    # Agrupar por división y contar procedimientos
    conteo_procedimientos = procedimientos.values(
        'id_division__division'  # Agrupar por nombre de la división
    ).annotate(count=Count('id')).order_by('id_division__division')

    return JsonResponse(list(conteo_procedimientos), safe=False)

# Api para generar valores para la grafica de procedimientos por tipo
def api_procedimientos_tipo_parroquias(request):
    tipo_procedimiento_id = request.GET.get('tipo_procedimiento_id')
    mes = request.GET.get('mes')

    # Filtrar por tipo de procedimiento
    procedimientos = Procedimientos.objects.all()
    if tipo_procedimiento_id:
        procedimientos = procedimientos.filter(id_tipo_procedimiento=tipo_procedimiento_id)

    # Filtrar por mes si se proporciona
    if mes:
        fecha_inicio = datetime.strptime(mes, '%Y-%m').date()
        fecha_fin = fecha_inicio.replace(day=1) + relativedelta(months=1)
        procedimientos = procedimientos.filter(fecha__gte=fecha_inicio, fecha__lt=fecha_fin)

    # Agrupar por parroquia y contar procedimientos
    conteo_procedimientos = procedimientos.values(
        'id_parroquia__parroquia'  # Cambia esto si el campo se llama de otra manera
    ).annotate(count=Count('id')).order_by('id_parroquia__parroquia')  # Ordenar por nombre de la parroquia

    return JsonResponse(list(conteo_procedimientos), safe=False)

# Api procedimiento por tipo de servicio - Tipo de detalles
# API para generar valores para la gráfica de procedimientos por tipo y detalles específicos
def api_procedimientos_tipo_detalles(request):
    tipo_procedimiento_id = request.GET.get('tipo_procedimiento_id')
    mes = request.GET.get('mes')

    # Filtrar procedimientos por tipo de procedimiento y mes
    procedimientos = Procedimientos.objects.all()
    if tipo_procedimiento_id:
        procedimientos = procedimientos.filter(id_tipo_procedimiento=tipo_procedimiento_id)

    if mes:
        fecha_inicio = datetime.strptime(mes, '%Y-%m').date()
        fecha_fin = fecha_inicio.replace(day=1) + relativedelta(months=1)
        procedimientos = procedimientos.filter(fecha__gte=fecha_inicio, fecha__lt=fecha_fin)

    # Variable para almacenar los resultados
    resultados = []

    # Filtrar y agrupar según el tipo de procedimiento
    if tipo_procedimiento_id == "1":  # Ejemplo: ID para "Abastecimiento de Agua"
        abastecimiento_data = Abastecimiento_agua.objects.filter(id_procedimiento__in=procedimientos).values(
            'id_tipo_servicio__nombre_institucion'  # Asume que `nombre` es el campo en Tipo_Institucion
        ).annotate(count=Count('id')).order_by('id_tipo_servicio__nombre_institucion')

        resultados = [
            {"tipo_servicio": item['id_tipo_servicio__nombre_institucion'], "count": item['count']}
            for item in abastecimiento_data
        ]
    # Filtrar y agrupar según el tipo de procedimiento
    if tipo_procedimiento_id == "2":  # Ejemplo: ID para "Abastecimiento de Agua"
        abastecimiento_data = Apoyo_Unidades.objects.filter(id_procedimiento__in=procedimientos).values(
            'id_tipo_apoyo__tipo_apoyo'  # Asume que `nombre` es el campo en Tipo_Institucion
        ).annotate(count=Count('id')).order_by('id_tipo_apoyo__tipo_apoyo')

        resultados = [
            {"tipo_servicio": item['id_tipo_apoyo__tipo_apoyo'], "count": item['count']}
            for item in abastecimiento_data
        ]
    # Filtrar y agrupar según el tipo de procedimiento
    if tipo_procedimiento_id == "3":  # Ejemplo: ID para "Abastecimiento de Agua"
        abastecimiento_data = Guardia_prevencion.objects.filter(id_procedimiento__in=procedimientos).values(
            'id_motivo_prevencion__motivo'  # Asume que `nombre` es el campo en Tipo_Institucion
        ).annotate(count=Count('id')).order_by('id_motivo_prevencion__motivo')

        resultados = [
            {"tipo_servicio": item['id_motivo_prevencion__motivo'], "count": item['count']}
            for item in abastecimiento_data
        ]

    # Filtrar y agrupar según el tipo de procedimiento
    if tipo_procedimiento_id == "5":  # Ejemplo: ID para "Abastecimiento de Agua"
        abastecimiento_data = Despliegue_Seguridad.objects.filter(id_procedimiento__in=procedimientos).values(
            'motivo_despliegue__motivo'  # Asume que `nombre` es el campo en Tipo_Institucion
        ).annotate(count=Count('id')).order_by('motivo_despliegue__motivo')

        resultados = [
            {"tipo_servicio": item['motivo_despliegue__motivo'], "count": item['count']}
            for item in abastecimiento_data
        ]
    
    # Filtrar y agrupar según el tipo de procedimiento
    if tipo_procedimiento_id == "12":  # Ejemplo: ID para "Abastecimiento de Agua"
        abastecimiento_data = Fallecidos.objects.filter(id_procedimiento__in=procedimientos).values(
            'motivo_fallecimiento'  # Asume que `nombre` es el campo en Tipo_Institucion
        ).annotate(count=Count('id')).order_by('motivo_fallecimiento')

        resultados = [
            {"tipo_servicio": item['motivo_fallecimiento'], "count": item['count']}
            for item in abastecimiento_data
        ]
    
    # Filtrar y agrupar según el tipo de procedimiento
    if tipo_procedimiento_id == "6":  # Ejemplo: ID para "Abastecimiento de Agua"
        abastecimiento_data = Falsa_Alarma.objects.filter(id_procedimiento__in=procedimientos).values(
            'motivo_alarma__motivo'  # Asume que `nombre` es el campo en Tipo_Institucion
        ).annotate(count=Count('id')).order_by('motivo_alarma__motivo')

        resultados = [
            {"tipo_servicio": item['motivo_alarma__motivo'], "count": item['count']}
            for item in abastecimiento_data
        ]
    
    # Filtrar y agrupar según el tipo de procedimiento
    if tipo_procedimiento_id == "9":  # Ejemplo: ID para "Abastecimiento de Agua"
        abastecimiento_data = Servicios_Especiales.objects.filter(id_procedimientos__in=procedimientos).values(
            'tipo_servicio__serv_especiales'  # Asume que `nombre` es el campo en Tipo_Institucion
        ).annotate(count=Count('id')).order_by('tipo_servicio__serv_especiales')

        resultados = [
            {"tipo_servicio": item['tipo_servicio__serv_especiales'], "count": item['count']}
            for item in abastecimiento_data
        ]

    # Filtrar y agrupar según el tipo de procedimiento
    if tipo_procedimiento_id == "10":  # Ejemplo: ID para "Abastecimiento de Agua"
        abastecimiento_data = Rescate.objects.filter(id_procedimientos__in=procedimientos).values(
            'tipo_rescate__tipo_rescate'  # Asume que `nombre` es el campo en Tipo_Institucion
        ).annotate(count=Count('id')).order_by('tipo_rescate__tipo_rescate')

        resultados = [
            {"tipo_servicio": item['tipo_rescate__tipo_rescate'], "count": item['count']}
            for item in abastecimiento_data
        ]

    # Filtrar y agrupar según el tipo de procedimiento
    if tipo_procedimiento_id == "11":  # Ejemplo: ID para "Abastecimiento de Agua"
        abastecimiento_data = Incendios.objects.filter(id_procedimientos__in=procedimientos).values(
            'id_tipo_incendio__tipo_incendio'  # Asume que `nombre` es el campo en Tipo_Institucion
        ).annotate(count=Count('id')).order_by('id_tipo_incendio__tipo_incendio')

        resultados = [
            {"tipo_servicio": item['id_tipo_incendio__tipo_incendio'], "count": item['count']}
            for item in abastecimiento_data
        ]

    # Filtrar y agrupar según el tipo de procedimiento
    if tipo_procedimiento_id == "7":  # Ejemplo: ID para "Abastecimiento de Agua"
        abastecimiento_data = Atenciones_Paramedicas.objects.filter(id_procedimientos__in=procedimientos).values(
            'tipo_atencion'  # Asume que `nombre` es el campo en Tipo_Institucion
        ).annotate(count=Count('id')).order_by('tipo_atencion')

        resultados = [
            {"tipo_servicio": item['tipo_atencion'], "count": item['count']}
            for item in abastecimiento_data
        ]

    # Filtrar y agrupar según el tipo de procedimiento
    if tipo_procedimiento_id == "16":  # Ejemplo: ID para "Abastecimiento de Agua"
        abastecimiento_data = Traslado_Prehospitalaria.objects.filter(id_procedimiento__in=procedimientos).values(
            'id_tipo_traslado__tipo_traslado'  # Asume que `nombre` es el campo en Tipo_Institucion
        ).annotate(count=Count('id')).order_by('id_tipo_traslado__tipo_traslado')

        resultados = [
            {"tipo_servicio": item['id_tipo_traslado__tipo_traslado'], "count": item['count']}
            for item in abastecimiento_data
        ]

    # Filtrar y agrupar según el tipo de procedimiento
    if tipo_procedimiento_id == "14":  # Ejemplo: ID para "Abastecimiento de Agua"
        abastecimiento_data = Evaluacion_Riesgo.objects.filter(id_procedimientos__in=procedimientos).values(
            'id_tipo_riesgo__tipo_riesgo'  # Asume que `nombre` es el campo en Tipo_Institucion
        ).annotate(count=Count('id')).order_by('id_tipo_riesgo__tipo_riesgo')

        resultados = [
            {"tipo_servicio": item['id_tipo_riesgo__tipo_riesgo'], "count": item['count']}
            for item in abastecimiento_data
        ]

    # Filtrar y agrupar según el tipo de procedimiento
    if tipo_procedimiento_id == "13":  # Ejemplo: ID para "Abastecimiento de Agua"
        abastecimiento_data = Mitigacion_Riesgos.objects.filter(id_procedimientos__in=procedimientos).values(
            'id_tipo_servicio__tipo_servicio'  # Asume que `nombre` es el campo en Tipo_Institucion
        ).annotate(count=Count('id')).order_by('id_tipo_servicio__tipo_servicio')

        resultados = [
            {"tipo_servicio": item['id_tipo_servicio__tipo_servicio'], "count": item['count']}
            for item in abastecimiento_data
        ]

    # Filtrar y agrupar según el tipo de procedimiento
    if tipo_procedimiento_id == "15":  # Ejemplo: ID para "Abastecimiento de Agua"
        abastecimiento_data = Puesto_Avanzada.objects.filter(id_procedimientos__in=procedimientos).values(
            'id_tipo_servicio__tipo_servicio'  # Asume que `nombre` es el campo en Tipo_Institucion
        ).annotate(count=Count('id')).order_by('id_tipo_servicio__tipo_servicio')

        resultados = [
            {"tipo_servicio": item['id_tipo_servicio__tipo_servicio'], "count": item['count']}
            for item in abastecimiento_data
        ]

    # Filtrar y agrupar según el tipo de procedimiento
    if tipo_procedimiento_id == "21":  # Ejemplo: ID para "Abastecimiento de Agua"
        abastecimiento_data = Retencion_Preventiva.objects.filter(id_procedimiento__in=procedimientos).values(
            'tipo_cilindro'  # Asume que `nombre` es el campo en Tipo_Institucion
        ).annotate(count=Count('id')).order_by('tipo_cilindro')

        resultados = [
            {"tipo_servicio": item['tipo_cilindro'], "count": item['count']}
            for item in abastecimiento_data
        ]

    # Filtrar y agrupar según el tipo de procedimiento
    if tipo_procedimiento_id == "22":  # Ejemplo: ID para "Abastecimiento de Agua"
        abastecimiento_data = Artificios_Pirotecnicos.objects.filter(id_procedimiento__in=procedimientos).values(
            'tipo_procedimiento__tipo'  # Asume que `nombre` es el campo en Tipo_Institucion
        ).annotate(count=Count('id')).order_by('tipo_procedimiento__tipo')

        resultados = [
            {"tipo_servicio": item['tipo_procedimiento__tipo'], "count": item['count']}
            for item in abastecimiento_data
        ]

    # Filtrar y agrupar según el tipo de procedimiento
    if tipo_procedimiento_id == "45":  # Ejemplo: ID para "Abastecimiento de Agua"
        abastecimiento_data = Procedimientos_Capacitacion.objects.filter(id_procedimientos__in=procedimientos).values(
            'tipo_capacitacion'  # Asume que `nombre` es el campo en Tipo_Institucion
        ).annotate(count=Count('id')).order_by('tipo_capacitacion')

        resultados = [
            {"tipo_servicio": item['tipo_capacitacion'], "count": item['count']}
            for item in abastecimiento_data
        ]

    # Filtrar y agrupar según el tipo de procedimiento
    if tipo_procedimiento_id == "18":  # Ejemplo: ID para "Abastecimiento de Agua"
        abastecimiento_data = Inspeccion_Prevencion_Asesorias_Tecnicas.objects.filter(id_procedimientos__in=procedimientos).values(
            'tipo_inspeccion'  # Asume que `nombre` es el campo en Tipo_Institucion
        ).annotate(count=Count('id')).order_by('tipo_inspeccion')

        resultados = [
            {"tipo_servicio": item['tipo_inspeccion'], "count": item['count']}
            for item in abastecimiento_data
        ]
        
        abastecimiento_data = Inspeccion_Habitabilidad.objects.filter(id_procedimientos__in=procedimientos).values(
            'tipo_inspeccion'  # Asume que `nombre` es el campo en Tipo_Institucion
        ).annotate(count=Count('id')).order_by('tipo_inspeccion')

        resultados += [
            {"tipo_servicio": item['tipo_inspeccion'], "count": item['count']}
            for item in abastecimiento_data
        ]

        abastecimiento_data = Inspeccion_Otros.objects.filter(id_procedimientos__in=procedimientos).values(
            'tipo_inspeccion'  # Asume que `nombre` es el campo en Tipo_Institucion
        ).annotate(count=Count('id')).order_by('tipo_inspeccion')

        resultados += [
            {"tipo_servicio": item['tipo_inspeccion'], "count": item['count']}
            for item in abastecimiento_data
        ]
        
        abastecimiento_data = Inspeccion_Arbol.objects.filter(id_procedimientos__in=procedimientos).values(
            'tipo_inspeccion'  # Asume que `nombre` es el campo en Tipo_Institucion
        ).annotate(count=Count('id')).order_by('tipo_inspeccion')

        resultados += [
            {"tipo_servicio": item['tipo_inspeccion'], "count": item['count']}
            for item in abastecimiento_data
        ]

    
    # Filtrar y agrupar según el tipo de procedimiento
    if tipo_procedimiento_id == "19":  # Ejemplo: ID para "Abastecimiento de Agua"
        abastecimiento_data = Investigacion.objects.filter(id_procedimientos__in=procedimientos).values(
            'tipo_siniestro'  # Asume que `nombre` es el campo en Tipo_Institucion
        ).annotate(count=Count('id')).order_by('tipo_siniestro')

        resultados = [
            {"tipo_servicio": item['tipo_siniestro'], "count": item['count']}
            for item in abastecimiento_data
        ]

    # Retornar el JSON con los resultados formateados
    return JsonResponse(resultados, safe=False)

# Api para generar los valores para la grafica de barras de la seccion de estadistica
def obtener_divisiones_estadistica(request):
    # Obtener el parámetro 'mes' (en formato 'YYYY-MM')
    mes = request.GET.get('mes', None)
    
    # Obtener el mes o el año actual si 'mes' no se proporciona
    hoy = datetime.now()
    if mes:
        try:
            # Si se proporciona el mes, obtenemos el primer día del mes
            primer_dia_mes = datetime.strptime(mes, '%Y-%m').date()
        except ValueError:
            return JsonResponse({"error": "Formato de mes inválido. Debe ser 'YYYY-MM'."}, status=400)
        ultimo_dia_mes = primer_dia_mes.replace(day=1).replace(month=primer_dia_mes.month % 12 + 1) - timedelta(days=1)
    else:
        # Si no se selecciona ningún mes, usar el año actual
        primer_dia_mes = hoy.replace(month=1, day=1).date()
        ultimo_dia_mes = hoy.replace(month=12, day=31).date()

    # Filtrar procedimientos por división en el rango de fechas determinado
    divisiones = {
        "rescate": {
            "total": Procedimientos.objects.filter(id_division=1, fecha__range=(primer_dia_mes, ultimo_dia_mes)).count(),
        },
        "operaciones": {
            "total": Procedimientos.objects.filter(id_division=2, fecha__range=(primer_dia_mes, ultimo_dia_mes)).count(),
        },
        "prevencion": {
            "total": Procedimientos.objects.filter(id_division=3, fecha__range=(primer_dia_mes, ultimo_dia_mes)).count(),
        },
        "grumae": {
            "total": Procedimientos.objects.filter(id_division=4, fecha__range=(primer_dia_mes, ultimo_dia_mes)).count(),
        },
        "prehospitalaria": {
            "total": Procedimientos.objects.filter(id_division=5, fecha__range=(primer_dia_mes, ultimo_dia_mes)).count(),
        },
        "enfermeria": {
            "total": Procedimientos.objects.filter(id_division=6, fecha__range=(primer_dia_mes, ultimo_dia_mes)).count(),
        },
        "servicios_medicos": {
            "total": Procedimientos.objects.filter(id_division=7, fecha__range=(primer_dia_mes, ultimo_dia_mes)).count(),
        },
        "psicologia": {
            "total": Procedimientos.objects.filter(id_division=8, fecha__range=(primer_dia_mes, ultimo_dia_mes)).count(),
        },
        "capacitacion": {
            "total": Procedimientos.objects.filter(id_division=9, fecha__range=(primer_dia_mes, ultimo_dia_mes)).count(),
        },
    }

    return JsonResponse(divisiones)

# Api para obtener valores para las cards de divisones de la seccion del dashboard
def obtener_divisiones(request):
    hoy = datetime.now().date()
    primer_dia_mes = hoy.replace(day=1)

    # Filtrado de procedimientos por división
    divisiones = {
        "rescate": {
            "total": Procedimientos.objects.filter(id_division=1).count(),
            "del_mes": Procedimientos.objects.filter(id_division=1, fecha__gte=primer_dia_mes).count(),
            "hoy": Procedimientos.objects.filter(id_division=1, fecha=hoy).count(),
        },
        "operaciones": {
            "total": Procedimientos.objects.filter(id_division=2).count(),
            "del_mes": Procedimientos.objects.filter(id_division=2, fecha__gte=primer_dia_mes).count(),
            "hoy": Procedimientos.objects.filter(id_division=2, fecha=hoy).count(),
        },
        "prevencion": {
            "total": Procedimientos.objects.filter(id_division=3).count(),
            "del_mes": Procedimientos.objects.filter(id_division=3, fecha__gte=primer_dia_mes).count(),
            "hoy": Procedimientos.objects.filter(id_division=3, fecha=hoy).count(),
        },
        "grumae": {
            "total": Procedimientos.objects.filter(id_division=4).count(),
            "del_mes": Procedimientos.objects.filter(id_division=4, fecha__gte=primer_dia_mes).count(),
            "hoy": Procedimientos.objects.filter(id_division=4, fecha=hoy).count(),
        },
        "prehospitalaria": {
            "total": Procedimientos.objects.filter(id_division=5).count(),
            "del_mes": Procedimientos.objects.filter(id_division=5, fecha__gte=primer_dia_mes).count(),
            "hoy": Procedimientos.objects.filter(id_division=5, fecha=hoy).count(),
        },
        "enfermeria": {
            "total": Procedimientos.objects.filter(id_division=6).count(),
            "del_mes": Procedimientos.objects.filter(id_division=6, fecha__gte=primer_dia_mes).count(),
            "hoy": Procedimientos.objects.filter(id_division=6, fecha=hoy).count(),
        },
        "servicios_medicos": {
            "total": Procedimientos.objects.filter(id_division=7).count(),
            "del_mes": Procedimientos.objects.filter(id_division=7, fecha__gte=primer_dia_mes).count(),
            "hoy": Procedimientos.objects.filter(id_division=7, fecha=hoy).count(),
        },
        "psicologia": {
            "total": Procedimientos.objects.filter(id_division=8).count(),
            "del_mes": Procedimientos.objects.filter(id_division=8, fecha__gte=primer_dia_mes).count(),
            "hoy": Procedimientos.objects.filter(id_division=8, fecha=hoy).count(),
        },
        "capacitacion": {
            "total": Procedimientos.objects.filter(id_division=9).count(),
            "del_mes": Procedimientos.objects.filter(id_division=9, fecha__gte=primer_dia_mes).count(),
            "hoy": Procedimientos.objects.filter(id_division=9, fecha=hoy).count(),
        },
    }

    return JsonResponse(divisiones)

# Api para Editar la informacion seleccionada del Personal
def edit_personal(request):
    if request.method == 'POST':
        # Obtener el ID de la persona
        persona_id = request.POST.get('personal_id')
        
        # Verificar si el ID está presente
        if persona_id:
            # Obtener el objeto Personal por su ID
            personal = get_object_or_404(Personal, id=persona_id)
            
            cedula = request.POST.get('formulario-cedula')
            nac = request.POST.get('formulario-nacionalidad')

            # Actualizar los campos del modelo
            personal.nombres = request.POST.get('formulario-nombres')
            personal.apellidos = request.POST.get('formulario-apellidos')
            personal.jerarquia = request.POST.get('formulario-jerarquia')
            personal.cargo = request.POST.get('formulario-cargo')
            personal.cedula = f"{nac}- {cedula}"
            personal.sexo = request.POST.get('formulario-sexo')
            personal.rol = request.POST.get('formulario-rol')
            personal.status = request.POST.get('formulario-status')
            
            # Guardar los cambios
            personal.save()
            
            # Redirigir o mostrar un mensaje de éxito
            return redirect('/personal/')  # Reemplaza con tu vista deseada
        else:
            # Si no se pasó el ID
            return redirect('/personal/')  # Redirigir a una página de error o donde desees
    
    # Si es GET, mostrar el formulario vacío o con los datos del objeto
    return render(request, 'editar_personal.html')

# Api para obtener el valor de el personal seleccionado
def get_persona(request, persona_id):
    try:
        persona = Personal.objects.get(id=persona_id)
        data = {
            'nombre': persona.nombres,
            'apellido': persona.apellidos,
            'cedula': persona.cedula,
            'jerarquia': persona.jerarquia,
            'cargo': persona.cargo,
            'rol': persona.rol,
            'sexo': persona.sexo,
            'status': persona.status,
        }
        return JsonResponse(data)
    except Personal.DoesNotExist:
        return JsonResponse({'error': 'Persona no encontrada'}, status=404)

# Login required
def login_required(view_func):
    def _wrapped_view(request, *args, **kwargs):
        if 'user' not in request.session:
            return redirect('/login/')  # Redirigir a la página de inicio de sesión
        return view_func(request, *args, **kwargs)
    return _wrapped_view

# Cierrre de sesion
def logout(request):
    request.session.flush()  # Eliminar todos los datos de la sesión
    return redirect('/login/')

def inicio(request):
    return render(request, 'blog/inicio.html')

def information(request):
    return render(request, 'blog/informacion.html')

# Vista de la Ventana Inicial (Login)
@never_cache
def Home(request):
    if request.method == "GET":
        return render(request, "index.html")
    else:
        usuario = request.POST["user"]
        contrasena = request.POST["password"]
        try:
            user = Usuarios.objects.get(user=usuario, password=contrasena)
            encargado = user.encargado  # Obtener el encargado relacionado
            # Guardar datos en la sesión
            request.session['user'] = {
                "user": user.user,
                "jerarquia": encargado.jerarquia,
                "nombres": encargado.nombres,
                "apellidos": encargado.apellidos,
            }
            return redirect("/dashboard/")
        except Usuarios.DoesNotExist:
            messages.error(request, 'Usuario o contraseña incorrectos')
            return render(request, 'index.html', {'error': True})

@login_required 
def View_personal(request):
    user = request.session.get('user')

    if not user:
        return redirect('/')
    
    personal = Personal.objects.exclude(id__in=[0, 4])
    personal = personal.order_by("id")
    # Lista de jerarquías en el orden deseado
    jerarquias = [
        "General", "Coronel", "Teniente Coronel", "Mayor", "Capitán", "Primer Teniente", 
        "Teniente", "Sargento Mayor", "Sargento Primero", "Sargento segundo", 
        "Cabo Primero", "Cabo Segundo", "Distinguido", "Bombero"
    ]

    # Filtro y ordenación de acuerdo a las jerarquías
    personal_ordenado =personal.order_by(
        Case(*[When(jerarquia=nombre, then=pos) for pos, nombre in enumerate(jerarquias)])
    )
    personal_count = personal_ordenado.count()

    if request.method == 'POST':
        formulario = FormularioRegistroPersonal(request.POST, prefix='formulario')

        if formulario.is_valid():
            
            new_personal = Personal(
                nombres = formulario.cleaned_data["nombres"],
                apellidos = formulario.cleaned_data["apellidos"],
                jerarquia = formulario.cleaned_data["jerarquia"],
                cargo = formulario.cleaned_data["cargo"],
                cedula = f"{formulario.cleaned_data['nacionalidad']}- {formulario.cleaned_data['cedula']}",
                sexo = formulario.cleaned_data["sexo"],
                rol = formulario.cleaned_data["rol"],
                status = formulario.cleaned_data["status"],
            )

            new_personal.save()

            return redirect("/personal/")

    else:
        formulario = FormularioRegistroPersonal(prefix='formulario')
    

    # Renderizar la página con los datos
    return render(request, "personal.html", {
        "user": user,
        "jerarquia": user["jerarquia"],
        "nombres": user["nombres"],
        "apellidos": user["apellidos"],
        "form_personal": formulario,
        "personal": personal_ordenado,
        "total": personal_count
    })

@login_required
def Dashboard(request):
    user = request.session.get('user')

    if not user:
        return redirect('/')
    # Renderizar la página con los datos
    return render(request, "dashboard.html", {
        "user": user,
        "jerarquia": user["jerarquia"],
        "nombres": user["nombres"],
        "apellidos": user["apellidos"],
    })

@login_required
# Vista de archivo para hacer pruebas de backend
def Prueba(request):

    if request.method == 'POST':
        form = Selecc_Tipo_Procedimiento(request.POST)
        if form.is_valid():
            usuarios = Usuarios.objects.all()
            divisiones = Divisiones.objects.all()
            procedimientos = Procedimientos.objects.all()

            valor_seleccionado = form.cleaned_data['tipo_procedimiento']
            # Aquí puedes hacer algo con el valor seleccionado
            return render(request, "prueba.html", {
            "usuarios": usuarios,
            "divisiones": divisiones,
            "procedimientos": procedimientos,
            "form": Formulario_Incendio(),
            })
    else:
        form = Selecc_Tipo_Procedimiento(),

        usuarios = Usuarios.objects.all()
        divisiones = Divisiones.objects.all()
        procedimientos = Procedimientos.objects.all()

        return render(request, "prueba.html", {
            "usuarios": usuarios,
            "divisiones": divisiones,
            "procedimientos": procedimientos,
            "form": Formulario_Traslado_Accidente(),
            })

# Vista de archivo para hacer pruebas de backend
def View_Procedimiento(request):
    user = request.session.get('user')
    if not user:
        return redirect('/')

    result = None

    if request.method == 'POST':
        form = SelectorDivision(request.POST, prefix='form1')
        form2 = SeleccionarInfo(request.POST, prefix='form2')
        form3 = Datos_Ubicacion(request.POST, prefix='form3')
        form4 = Selecc_Tipo_Procedimiento(request.POST, prefix='form4')
        abast_agua = formulario_abastecimiento_agua(request.POST, prefix='abast_agua')
        apoyo_unid = Formulario_apoyo_unidades(request.POST, prefix='apoyo_unid')
        guard_prev = Formulario_guardia_prevencion(request.POST, prefix='guard_prev')
        atend_no_efec = Formulario_atendido_no_efectuado(request.POST, prefix='atend_no_efec')
        desp_seguridad = Formulario_despliegue_seguridad(request.POST, prefix='desp_seguridad')
        fals_alarm = Formulario_falsa_alarma(request.POST, prefix='fals_alarm')
        serv_especial = Formulario_Servicios_Especiales(request.POST, prefix='serv_especial')
        form_fallecido = Formulario_Fallecidos(request.POST, prefix='form_fallecido')
        rescate_form = Formulario_Rescate(request.POST, prefix='rescate_form')
        incendio_form = Formulario_Incendio(request.POST, prefix='incendio_form')
        atenciones_paramedicas = Formulario_Atenciones_Paramedicas(request.POST, prefix='atenciones_paramedicas')

        emergencias_medicas = Formulario_Emergencias_Medicas(request.POST, prefix='emergencias_medicas')
        traslados_emergencias = Formulario_Traslados(request.POST, prefix='traslados_emergencias')

        persona_presente_form = Formulario_Persona_Presente(request.POST, prefix='persona_presente_form')
        detalles_vehiculo_form = Formulario_Detalles_Vehiculos(request.POST, prefix='detalles_vehiculo_form')

        formulario_accidentes_transito = Formulario_Accidentes_Transito(request.POST, prefix='formulario_accidentes_transito')
        detalles_lesionados_accidentes = Formulario_Detalles_Lesionados(request.POST, prefix='detalles_lesionados_accidentes')
        detalles_lesionados_accidentes2 = Formulario_Detalles_Lesionados2(request.POST, prefix='detalles_lesionados_accidentes2')
        detalles_lesionados_accidentes3 = Formulario_Detalles_Lesionados3(request.POST, prefix='detalles_lesionados_accidentes3')
        traslados_accidentes = Formulario_Traslado_Accidente(request.POST, prefix='traslados_accidentes')
        traslados_accidentes2 = Formulario_Traslado_Accidente2(request.POST, prefix='traslados_accidentes2')
        traslados_accidentes3 = Formulario_Traslado_Accidente3(request.POST, prefix='traslados_accidentes3')
        detalles_vehiculo_accidentes = Formulario_Detalles_Vehiculos(request.POST, prefix='detalles_vehiculos_accidentes')
        detalles_vehiculo_accidentes2 = Formulario_Detalles_Vehiculos2(request.POST, prefix='detalles_vehiculos_accidentes2')
        detalles_vehiculo_accidentes3 = Formulario_Detalles_Vehiculos3(request.POST, prefix='detalles_vehiculos_accidentes3')

        rescate_form_persona = Formulario_Rescate_Persona(request.POST, prefix='rescate_form_persona')
        rescate_form_animal = Formulario_Rescate_Animal(request.POST, prefix='rescate_form_animal')

        evaluacion_riesgo_form = Forulario_Evaluacion_Riesgo(request.POST, prefix='evaluacion_riesgo_form')
        mitigacion_riesgo_form = Formulario_Mitigacion_Riesgos(request.POST, prefix='mitigacion_riesgo_form')
        puesto_avanzada_form = Formulario_Puesto_Avanzada(request.POST, prefix='puesto_avanzada_form')
        traslados_prehospitalaria_form = Formulario_Traslados_Prehospitalaria(request.POST, prefix='traslados_prehospitalaria_form')
        asesoramiento_form = Formulario_Asesoramiento(request.POST, prefix='asesoramiento_form')
        persona_presente_eval_form = Formularia_Persona_Presente_Eval(request.POST, prefix='persona_presente_eval_form')
        reinspeccion_prevencion = Formulario_Reinspeccion_Prevencion(request.POST, prefix='reinspeccion_prevencion')
        retencion_preventiva = Formulario_Retencion_Preventiva(request.POST, prefix='retencion_preventiva')

        artificios_pirotecnico = Formulario_Artificios_Pirotecnicos(request.POST, prefix='artificios_pirotecnico')
        lesionados = Formulario_Lesionado(request.POST, prefix='lesionados')
        incendio_art = Formulario_Incendio_Art(request.POST, prefix='incendio_art')
        persona_presente_art = Formulario_Persona_Presente_Art(request.POST, prefix='persona_presente_art')
        detalles_vehiculo_art = Formulario_Detalles_Vehiculos_Incendio_Art(request.POST, prefix='detalles_vehiculo_art')
        fallecidos_art = Formulario_Fallecidos_Art(request.POST, prefix='fallecidos_art')
        inspeccion_artificios_pir = Formulario_Inspeccion_Establecimiento_Art(request.POST, prefix='inspeccion_artificios_pir')
        form_enfermeria = Formulario_Enfermeria(request.POST, prefix='form_enfermeria')
        servicios_medicos = Formulario_Servicios_medicos(request.POST, prefix='form_servicios_medicos')
        psicologia = Formulario_psicologia(request.POST,prefix='form_psicologia')
        capacitacion = Formulario_capacitacion(request.POST,prefix='form_capacitacion')
        form_valoracion_medica = Formulario_Valoracion_Medica(request.POST, prefix='form_valoracion_medica')
        form_detalles_enfermeria = Formulario_Detalles_Enfermeria(request.POST, prefix='form_detalles_enfermeria')
        form_detalles_psicologia = Formulario_Procedimientos_Psicologia(request.POST, prefix='form_detalles_psicologia')
        
        form_capacitacion = Formulario_Capacitacion_Proc(request.POST,prefix='form_capacitacion')
        form_frente_preventivo = Formulario_Frente_Preventivo(request.POST,prefix='form_frente_preventivo')
        form_jornada_medica = Formulario_Jornada_Medica(request.POST, prefix='form_jornada_medica')

        form_inspecciones = Formulario_Inspecciones(request.POST, prefix='form_inspecciones')
        form_inspecciones_prevencion = Formulario_Inspeccion_Prevencion_Asesorias_Tecnicas(request.POST, prefix='form_inspecciones_prevencion')
        form_inspecciones_habitabilidad = Formulario_Inspeccion_Habitabilidad(request.POST, prefix='form_inspecciones_habitabilidad')
        form_inspecciones_arbol = Formulario_Inspeccion_Arbol(request.POST, prefix='form_inspecciones_arbol')
        form_inspecciones_otros = Formulario_Inspeccion_Otros(request.POST, prefix='form_inspecciones_otros')

        form_investigacion = Formulario_Investigacion(request.POST, prefix='form_investigacion')
        form_inv_vehiculo = Formulario_Investigacion_Vehiculo(request.POST, prefix='form_inv_vehiculo')
        form_inv_comercio = Formulario_Investigacion_Comercio(request.POST, prefix='form_inv_comercio')
        form_inv_estructura = Formulario_Investigacion_Estructura_Vivienda(request.POST, prefix='form_inv_estructura')
        # Imprimir request.POST para depuración

        if not form.is_valid():
            print("Errores en form1:", form.errors)
            result = True
        if not form2.is_valid():
            print("Errores en form2:", form2.errors)
            result = True
        if not form3.is_valid():
            print("Errores en form3:", form3.errors)
            result = True
        if not form4.is_valid():
            print("Errores en form4:", form4.errors)
            result = True

        if form.is_valid():
            result = False

            division = form.cleaned_data["opciones"]
            tipo_procedimiento = ""

            if (division == "1" or division == "2" or division == "3" or division == "4" or division == "5") and (form2.is_valid() and form3.is_valid() and form4.is_valid()):
                solicitante = form2.cleaned_data["solicitante"]
                solicitante_externo = form2.cleaned_data["solicitante_externo"]
                efectivos_enviados = form2.cleaned_data["efectivos_enviados"]
                jefe_comision = form2.cleaned_data["jefe_comision"]
                municipio = form3.cleaned_data["municipio"]
                direccion = form3.cleaned_data["direccion"]
                fecha = form3.cleaned_data["fecha"]
                hora = form3.cleaned_data["hora"]
                tipo_procedimiento = form4.cleaned_data["tipo_procedimiento"]
                parroquia = form3.cleaned_data["parroquia"]

                division_instance = Divisiones.objects.get(id=division)
                jefe_comision_instance = Personal.objects.get(id=jefe_comision)
                municipio_instance = Municipios.objects.get(id=municipio)
                tipo_procedimiento_instance = Tipos_Procedimientos.objects.get(id=tipo_procedimiento)

                if solicitante:
                    solicitante_instance = Personal.objects.get(id=solicitante)


                if solicitante_externo=="":
                    solicitante_externo = ""

                # # Crear una nueva instancia del modelo Procedimientos
                nuevo_procedimiento = Procedimientos(
                    id_division=division_instance,
                    id_solicitante=solicitante_instance,
                    solicitante_externo=solicitante_externo,
                    efectivos_enviados=efectivos_enviados,
                    id_jefe_comision=jefe_comision_instance,
                    id_municipio=municipio_instance,
                    direccion=direccion,
                    fecha=fecha,
                    hora=hora,
                    id_tipo_procedimiento=tipo_procedimiento_instance
                )

                # # Solo asignar parroquia si está presente
                if parroquia:
                    parroquia_instance = Parroquias.objects.get(id=parroquia)
                    nuevo_procedimiento.id_parroquia = parroquia_instance

                if division != "3":
                    unidad = form2.cleaned_data["unidad"]
                    unidad_instance = Unidades.objects.get(id=unidad)
                    nuevo_procedimiento.unidad=unidad_instance

                nuevo_procedimiento.save()

            if division == "6" and form_enfermeria.is_valid():
                dependencia = form_enfermeria.cleaned_data["dependencia"]
                encargado_area = form_enfermeria.cleaned_data["encargado_area"]
                municipio = form3.cleaned_data["municipio"]
                direccion = form3.cleaned_data["direccion"]
                fecha = form3.cleaned_data["fecha"]
                hora = form3.cleaned_data["hora"]
                tipo_procedimiento = form4.cleaned_data["tipo_procedimiento"]
                parroquia = form3.cleaned_data["parroquia"]

                division_instance = Divisiones.objects.get(id=division)
                municipio_instance = Municipios.objects.get(id=municipio)
                tipo_procedimiento_instance = Tipos_Procedimientos.objects.get(id=tipo_procedimiento)

                # # Crear una nueva instancia del modelo Procedimientos
                nuevo_procedimiento = Procedimientos(
                    id_division=division_instance,
                    dependencia=dependencia,
                    solicitante_externo=encargado_area,
                    id_municipio=municipio_instance,
                    direccion=direccion,
                    fecha=fecha,
                    hora=hora,
                    id_tipo_procedimiento=tipo_procedimiento_instance
                )

                # # Solo asignar parroquia si está presente
                if parroquia:
                    parroquia_instance = Parroquias.objects.get(id=parroquia)
                    nuevo_procedimiento.id_parroquia = parroquia_instance

                nuevo_procedimiento.save()
            
            if division == "7" and servicios_medicos.is_valid():
                tipo_servicio = servicios_medicos.cleaned_data["tipo_servicio"]
                jefe_area = servicios_medicos.cleaned_data["jefe_area"]
                municipio = form3.cleaned_data["municipio"]
                direccion = form3.cleaned_data["direccion"]
                fecha = form3.cleaned_data["fecha"]
                hora = form3.cleaned_data["hora"]
                tipo_procedimiento = form4.cleaned_data["tipo_procedimiento"]
                parroquia = form3.cleaned_data["parroquia"]

                division_instance = Divisiones.objects.get(id=division)
                municipio_instance = Municipios.objects.get(id=municipio)
                tipo_procedimiento_instance = Tipos_Procedimientos.objects.get(id=tipo_procedimiento)

                # # Crear una nueva instancia del modelo Procedimientos
                nuevo_procedimiento = Procedimientos(
                    id_division=division_instance,
                    tipo_servicio=tipo_servicio,
                    solicitante_externo=jefe_area,
                    id_municipio=municipio_instance,
                    direccion=direccion,
                    fecha=fecha,
                    hora=hora,
                    id_tipo_procedimiento=tipo_procedimiento_instance
                )

                # # Solo asignar parroquia si está presente
                if parroquia:
                    parroquia_instance = Parroquias.objects.get(id=parroquia)
                    nuevo_procedimiento.id_parroquia = parroquia_instance

                nuevo_procedimiento.save()

            if division == "8" and psicologia.is_valid():
                jefe_area = psicologia.cleaned_data["jefe_area"]
                municipio = form3.cleaned_data["municipio"]
                direccion = form3.cleaned_data["direccion"]
                fecha = form3.cleaned_data["fecha"]
                hora = form3.cleaned_data["hora"]
                tipo_procedimiento = form4.cleaned_data["tipo_procedimiento"]
                parroquia = form3.cleaned_data["parroquia"]

                division_instance = Divisiones.objects.get(id=division)
                municipio_instance = Municipios.objects.get(id=municipio)
                tipo_procedimiento_instance = Tipos_Procedimientos.objects.get(id=tipo_procedimiento)

                # # Crear una nueva instancia del modelo Procedimientos
                nuevo_procedimiento = Procedimientos(
                    id_division=division_instance,
                    solicitante_externo=jefe_area,
                    id_municipio=municipio_instance,
                    direccion=direccion,
                    fecha=fecha,
                    hora=hora,
                    id_tipo_procedimiento=tipo_procedimiento_instance
                )

                # # Solo asignar parroquia si está presente
                if parroquia:
                    parroquia_instance = Parroquias.objects.get(id=parroquia)
                    nuevo_procedimiento.id_parroquia = parroquia_instance

                nuevo_procedimiento.save()

            if division == "9" and capacitacion.is_valid():
                dependencia = capacitacion.cleaned_data["dependencia"]
                instructor = capacitacion.cleaned_data["instructor"]
                solicitante = capacitacion.cleaned_data["solicitante"]
                solicitante_externo = capacitacion.cleaned_data["solicitante_externo"]
                municipio = form3.cleaned_data["municipio"]
                direccion = form3.cleaned_data["direccion"]
                fecha = form3.cleaned_data["fecha"]
                hora = form3.cleaned_data["hora"]

                parroquia = form3.cleaned_data["parroquia"]
                tipo_procedimiento = 45

                division_instance = Divisiones.objects.get(id=division)
                municipio_instance = Municipios.objects.get(id=municipio)
                tipo_procedimiento_instance = Tipos_Procedimientos.objects.get(id=tipo_procedimiento)
                jefe_comision_instance = Personal.objects.get(id=instructor)

                if solicitante:
                    solicitante_instance = Personal.objects.get(id=solicitante)

                if solicitante_externo=="":
                    solicitante_externo = ""


                # # Crear una nueva instancia del modelo Procedimientos
                nuevo_procedimiento = Procedimientos(
                    id_division=division_instance,
                    dependencia=dependencia,
                    id_jefe_comision=jefe_comision_instance,
                    id_solicitante=solicitante_instance,
                    solicitante_externo=solicitante_externo,
                    id_municipio=municipio_instance,
                    direccion=direccion,
                    fecha=fecha,
                    hora=hora,
                    id_tipo_procedimiento=tipo_procedimiento_instance
                )

                # # Solo asignar parroquia si está presente
                if parroquia:
                    parroquia_instance = Parroquias.objects.get(id=parroquia)
                    nuevo_procedimiento.id_parroquia = parroquia_instance

                nuevo_procedimiento.save()

                if dependencia == "Capacitacion" and form_capacitacion.is_valid():
                    tipo_capacitacion = form_capacitacion.cleaned_data["tipo_capacitacion"]
                    tipo_clasificacion = form_capacitacion.cleaned_data["tipo_clasificacion"]
                    personas_beneficiadas = form_capacitacion.cleaned_data["personas_beneficiadas"]
                    descripcion = form_capacitacion.cleaned_data["descripcion"]
                    material_utilizado = form_capacitacion.cleaned_data["material_utilizado"]
                    status = form_capacitacion.cleaned_data["status"]

                    new_detalles_capacitacion = Procedimientos_Capacitacion(
                        id_procedimientos = nuevo_procedimiento,
                        tipo_capacitacion = tipo_capacitacion,
                        tipo_clasificacion = tipo_clasificacion,
                        personas_beneficiadas = personas_beneficiadas,
                        descripcion = descripcion,
                        material_utilizado = material_utilizado,
                        status = status
                    )

                    new_detalles_capacitacion.save()

                if dependencia == "Frente Preventivo" and form_frente_preventivo.is_valid():
                    nombre_actividad = form_frente_preventivo.cleaned_data["nombre_actividad"]
                    estrategia = form_frente_preventivo.cleaned_data["estrategia"]
                    personas_beneficiadas = form_frente_preventivo.cleaned_data["personas_beneficiadas"]
                    descripcion = form_frente_preventivo.cleaned_data["descripcion"]
                    material_utilizado = form_frente_preventivo.cleaned_data["material_utilizado"]
                    status = form_frente_preventivo.cleaned_data["status"]

                    new_detalles_frente_preventivo = Procedimientos_Frente_Preventivo(
                        id_procedimientos = nuevo_procedimiento,
                        nombre_actividad = nombre_actividad,
                        estrategia = estrategia,
                        personas_beneficiadas = personas_beneficiadas,
                        descripcion = descripcion,
                        material_utilizado = material_utilizado,
                        status = status
                    )

                    new_detalles_frente_preventivo.save()

            # Ahora dependiendo del tipo de procedimiento, verifica el formulario correspondiente y guarda la instancia
            if tipo_procedimiento == "1" and abast_agua.is_valid():
                # Abastecimiento de Agua
                nacionalidad=abast_agua.cleaned_data["nacionalidad"]
                cedula=abast_agua.cleaned_data["cedula"]

                nuevo_abast_agua = Abastecimiento_agua(
                    id_procedimiento=nuevo_procedimiento,
                    id_tipo_servicio=Tipo_Institucion.objects.get(id=abast_agua.cleaned_data["tipo_servicio"]),
                    nombres=abast_agua.cleaned_data["nombres"],
                    apellidos=abast_agua.cleaned_data["apellidos"],
                    cedula=f"{nacionalidad}-{cedula}",
                    ltrs_agua=abast_agua.cleaned_data["ltrs_agua"],
                    personas_atendidas=abast_agua.cleaned_data["personas_atendidas"],
                    descripcion=abast_agua.cleaned_data["descripcion"],
                    material_utilizado=abast_agua.cleaned_data["material_utilizado"],
                    status=abast_agua.cleaned_data["status"]
                )
                nuevo_abast_agua.save()

            if tipo_procedimiento == "2" and apoyo_unid.is_valid():
                tipo_apoyo = apoyo_unid.cleaned_data["tipo_apoyo"]
                unidad_apoyada = apoyo_unid.cleaned_data["unidad_apoyada"]
                descripcion = apoyo_unid.cleaned_data["descripcion"]
                material_utilizado = apoyo_unid.cleaned_data["material_utilizado"]
                status = apoyo_unid.cleaned_data["status"]

                tipo_apoyo_instance = Tipo_apoyo.objects.get(id=tipo_apoyo)

                nuevo_apoyo_unidad = Apoyo_Unidades(
                    id_procedimiento=nuevo_procedimiento,
                    id_tipo_apoyo=tipo_apoyo_instance,
                    unidad_apoyada=unidad_apoyada,
                    descripcion=descripcion,
                    material_utilizado=material_utilizado,
                    status=status
                )
                nuevo_apoyo_unidad.save()

            if tipo_procedimiento == "3" and guard_prev.is_valid():
                mot_prevencion = guard_prev.cleaned_data["motivo_prevencion"]
                descripcion = guard_prev.cleaned_data["descripcion"]
                material_utilizado = guard_prev.cleaned_data["material_utilizado"]
                status = guard_prev.cleaned_data["status"]

                Tipo_Motivo_instance = Motivo_Prevencion.objects.get(id=mot_prevencion)

                nuevo_guard_prevencion = Guardia_prevencion(
                    id_procedimiento=nuevo_procedimiento,
                    id_motivo_prevencion=Tipo_Motivo_instance,
                    descripcion=descripcion,
                    material_utilizado=material_utilizado,
                    status=status
                )
                nuevo_guard_prevencion.save()

            if tipo_procedimiento == "4" and atend_no_efec.is_valid():
                descripcion = atend_no_efec.cleaned_data["descripcion"]
                material_utilizado = atend_no_efec.cleaned_data["material_utilizado"]
                status = atend_no_efec.cleaned_data["status"]

                nuevo_atend_no_efect = Atendido_no_Efectuado(
                    id_procedimiento=nuevo_procedimiento,
                    descripcion=descripcion,
                    material_utilizado=material_utilizado,
                    status=status
                )
                nuevo_atend_no_efect.save()

            if tipo_procedimiento == "5" and desp_seguridad.is_valid():
                descripcion = desp_seguridad.cleaned_data["descripcion"]
                material_utilizado = desp_seguridad.cleaned_data["material_utilizado"]
                status =desp_seguridad.cleaned_data["status"]
                motv_despliegue = desp_seguridad.cleaned_data["motv_despliegue"]

                Tipo_Motivo_instance = Motivo_Despliegue.objects.get(id=motv_despliegue)

                desp_seguridad = Despliegue_Seguridad(
                    id_procedimiento=nuevo_procedimiento,
                    motivo_despliegue = Tipo_Motivo_instance,
                    descripcion=descripcion,
                    material_utilizado=material_utilizado,
                    status=status
                )
                desp_seguridad.save()

            if tipo_procedimiento == "6" and fals_alarm.is_valid():
                descripcion = fals_alarm.cleaned_data["descripcion"]
                material_utilizado = fals_alarm.cleaned_data["material_utilizado"]
                status = fals_alarm.cleaned_data["status"]
                motv_alarma = fals_alarm.cleaned_data["motv_alarma"]

                Tipo_Motivo_instance = Motivo_Alarma.objects.get(id=motv_alarma)

                nueva_falsa_alarma = Falsa_Alarma(
                    id_procedimiento=nuevo_procedimiento,
                    motivo_alarma = Tipo_Motivo_instance,
                    descripcion=descripcion,
                    material_utilizado=material_utilizado,
                    status=status
                )
                nueva_falsa_alarma.save()

            if tipo_procedimiento == "7" and atenciones_paramedicas.is_valid():

                tipo_atencion = atenciones_paramedicas.cleaned_data["tipo_atencion"]

                nueva_atencion_paramedica = Atenciones_Paramedicas(
                  id_procedimientos = nuevo_procedimiento,
                  tipo_atencion = tipo_atencion
                )
                nueva_atencion_paramedica.save()

                if tipo_atencion == "Emergencias Medicas" and emergencias_medicas.is_valid():
                    nombre = emergencias_medicas.cleaned_data["nombre"]
                    apellido = emergencias_medicas.cleaned_data["apellido"]
                    nacionalidad = emergencias_medicas.cleaned_data["nacionalidad"]
                    cedula = emergencias_medicas.cleaned_data["cedula"]
                    edad = emergencias_medicas.cleaned_data["edad"]
                    sexo = emergencias_medicas.cleaned_data["sexo"]
                    idx = emergencias_medicas.cleaned_data["idx"]
                    descripcion = emergencias_medicas.cleaned_data["descripcion"]
                    material_utilizado = emergencias_medicas.cleaned_data["material_utilizado"]
                    status = emergencias_medicas.cleaned_data["status"]
                    trasladado = emergencias_medicas.cleaned_data["trasladado"]

                    nueva_emergencia_medica = Emergencias_Medicas(
                       id_atencion = nueva_atencion_paramedica,
                       nombres = nombre,
                       apellidos = apellido,
                       cedula = f"{nacionalidad}-{cedula}",
                       edad = edad,
                       sexo = sexo,
                       idx = idx,
                       descripcion = descripcion,
                       material_utilizado = material_utilizado,
                       status = status,
                    )
                    nueva_emergencia_medica.save()

                    if trasladado == True and traslados_emergencias.is_valid():
                        hospital = traslados_emergencias.cleaned_data["hospital_trasladado"]
                        medico = traslados_emergencias.cleaned_data["medico_receptor"]
                        mpps_cmt = traslados_emergencias.cleaned_data["mpps_cmt"]

                        nuevo_traslado_emergencia = Traslado(
                           id_lesionado = nueva_emergencia_medica,
                           hospital_trasladado = hospital,
                           medico_receptor = medico,
                           mpps_cmt = mpps_cmt,
                        )
                        nuevo_traslado_emergencia.save()

                if tipo_atencion == "Accidentes de Transito" and formulario_accidentes_transito.is_valid():
                    tipo_accidente = formulario_accidentes_transito.cleaned_data["tipo_accidente"]
                    cantidad_lesionado = formulario_accidentes_transito.cleaned_data["cantidad_lesionado"]
                    material_utilizado = formulario_accidentes_transito.cleaned_data["material_utilizado"]
                    status = formulario_accidentes_transito.cleaned_data["status"]
                    agg_vehiculo = formulario_accidentes_transito.cleaned_data["agg_vehiculo"]
                    agg_lesionado = formulario_accidentes_transito.cleaned_data["agg_lesionado"]

                    tipo_accidente_instance = Tipo_Accidente.objects.get(id=tipo_accidente)

                    nuevo_accidente_transito = Accidentes_Transito(
                      id_atencion = nueva_atencion_paramedica,
                      tipo_de_accidente = tipo_accidente_instance,
                      cantidad_lesionados = cantidad_lesionado,
                      material_utilizado = material_utilizado,
                      status = status,
                    )
                    nuevo_accidente_transito.save()

                    if agg_vehiculo == True and detalles_vehiculo_accidentes.is_valid():
                        modelo1 = detalles_vehiculo_accidentes.cleaned_data["modelo"]
                        marca1 = detalles_vehiculo_accidentes.cleaned_data["marca"]
                        color1 = detalles_vehiculo_accidentes.cleaned_data["color"]
                        año1 = detalles_vehiculo_accidentes.cleaned_data["año"]
                        placas1 = detalles_vehiculo_accidentes.cleaned_data["placas"]
                        agg_vehiculo2 = detalles_vehiculo_accidentes.cleaned_data["agg_vehiculo"]

                        nuevo_vehiculo_accidente = Detalles_Vehiculos_Accidente(
                            id_vehiculo = nuevo_accidente_transito,
                            modelo = modelo1,
                            marca = marca1,
                            color = color1,
                            año = año1,
                            placas = placas1,
                        )
                        nuevo_vehiculo_accidente.save()

                        if agg_vehiculo2 == True and detalles_vehiculo_accidentes2.is_valid():
                            modelo2 = detalles_vehiculo_accidentes2.cleaned_data["modelo"]
                            marca2 = detalles_vehiculo_accidentes2.cleaned_data["marca"]
                            color2 = detalles_vehiculo_accidentes2.cleaned_data["color"]
                            año2 = detalles_vehiculo_accidentes2.cleaned_data["año"]
                            placas2 = detalles_vehiculo_accidentes2.cleaned_data["placas"]
                            agg_vehiculo3 = detalles_vehiculo_accidentes2.cleaned_data["agg_vehiculo"]

                            nuevo_vehiculo_accidente2 = Detalles_Vehiculos_Accidente(
                                id_vehiculo = nuevo_accidente_transito,
                                modelo = modelo2,
                                marca = marca2,
                                color = color2,
                                año = año2,
                                placas = placas2,
                            )
                            nuevo_vehiculo_accidente2.save()

                            if agg_vehiculo3 == True and detalles_vehiculo_accidentes3.is_valid():
                                modelo3 = detalles_vehiculo_accidentes3.cleaned_data["modelo"]
                                marca3 = detalles_vehiculo_accidentes3.cleaned_data["marca"]
                                color3 = detalles_vehiculo_accidentes3.cleaned_data["color"]
                                año3 = detalles_vehiculo_accidentes3.cleaned_data["año"]
                                placas3 = detalles_vehiculo_accidentes3.cleaned_data["placas"]

                                nuevo_vehiculo_accidente3 = Detalles_Vehiculos_Accidente(
                                    id_vehiculo = nuevo_accidente_transito,
                                    modelo = modelo3,
                                    marca = marca3,
                                    color = color3,
                                    año = año3,
                                    placas = placas3,
                                )
                                nuevo_vehiculo_accidente3.save()


                    if agg_lesionado == True and detalles_lesionados_accidentes.is_valid():
                        nombre = detalles_lesionados_accidentes.cleaned_data["nombre"]
                        apellido = detalles_lesionados_accidentes.cleaned_data["apellido"]
                        nacionalidad = detalles_lesionados_accidentes.cleaned_data["nacionalidad"]
                        cedula = detalles_lesionados_accidentes.cleaned_data["cedula"]
                        edad = detalles_lesionados_accidentes.cleaned_data["edad"]
                        sexo = detalles_lesionados_accidentes.cleaned_data["sexo"]
                        idx = detalles_lesionados_accidentes.cleaned_data["idx"]
                        descripcion = detalles_lesionados_accidentes.cleaned_data["descripcion"]
                        trasladado = detalles_lesionados_accidentes.cleaned_data["trasladado"]
                        otro_lesionado = detalles_lesionados_accidentes.cleaned_data["otro_lesionado"]

                        nuevo_lesionado = Lesionados(
                            id_accidente = nuevo_accidente_transito,
                            nombres = nombre,
                            apellidos = apellido,
                            cedula = f"{nacionalidad}-{cedula}",
                            edad = edad,
                            sexo = sexo,
                            idx = idx,
                            descripcion = descripcion,
                        )
                        nuevo_lesionado.save()

                        if trasladado == True and traslados_accidentes.is_valid():
                            hospital = traslados_accidentes.cleaned_data["hospital_trasladado"]
                            medico = traslados_accidentes.cleaned_data["medico_receptor"]
                            mpps_cmt = traslados_accidentes.cleaned_data["mpps_cmt"]

                            nuevo_traslado_accidente = Traslado_Accidente(
                                id_lesionado = nuevo_lesionado,
                                hospital_trasladado = hospital,
                                medico_receptor = medico,
                                mpps_cmt = mpps_cmt
                            )
                            nuevo_traslado_accidente.save()

                        if otro_lesionado == True and detalles_lesionados_accidentes2.is_valid():
                            nombre = detalles_lesionados_accidentes2.cleaned_data["nombre"]
                            apellido = detalles_lesionados_accidentes2.cleaned_data["apellido"]
                            nacionalidad = detalles_lesionados_accidentes2.cleaned_data["nacionalidad"]
                            cedula = detalles_lesionados_accidentes2.cleaned_data["cedula"]
                            edad = detalles_lesionados_accidentes2.cleaned_data["edad"]
                            sexo = detalles_lesionados_accidentes2.cleaned_data["sexo"]
                            idx = detalles_lesionados_accidentes2.cleaned_data["idx"]
                            descripcion = detalles_lesionados_accidentes2.cleaned_data["descripcion"]
                            trasladado = detalles_lesionados_accidentes2.cleaned_data["trasladado"]
                            otro_lesionado = detalles_lesionados_accidentes2.cleaned_data["otro_lesionado"]

                            nuevo_lesionado = Lesionados(
                                id_accidente = nuevo_accidente_transito,
                                nombres = nombre,
                                apellidos = apellido,
                                cedula = f"{nacionalidad}-{cedula}",
                                edad = edad,
                                sexo = sexo,
                                idx = idx,
                                descripcion = descripcion,
                            )
                            nuevo_lesionado.save()

                            if trasladado == True and traslados_accidentes2.is_valid():
                                hospital = traslados_accidentes2.cleaned_data["hospital_trasladado"]
                                medico = traslados_accidentes2.cleaned_data["medico_receptor"]
                                mpps_cmt = traslados_accidentes2.cleaned_data["mpps_cmt"]

                                nuevo_traslado_accidente = Traslado_Accidente(
                                    id_lesionado = nuevo_lesionado,
                                    hospital_trasladado = hospital,
                                    medico_receptor = medico,
                                    mpps_cmt = mpps_cmt
                                )
                                nuevo_traslado_accidente.save()

                            if otro_lesionado == True and detalles_lesionados_accidentes3.is_valid():
                                nombre = detalles_lesionados_accidentes3.cleaned_data["nombre"]
                                apellido = detalles_lesionados_accidentes3.cleaned_data["apellido"]
                                nacionalidad = detalles_lesionados_accidentes3.cleaned_data["nacionalidad"]
                                cedula = detalles_lesionados_accidentes3.cleaned_data["cedula"]
                                edad = detalles_lesionados_accidentes3.cleaned_data["edad"]
                                sexo = detalles_lesionados_accidentes3.cleaned_data["sexo"]
                                idx = detalles_lesionados_accidentes3.cleaned_data["idx"]
                                descripcion = detalles_lesionados_accidentes3.cleaned_data["descripcion"]
                                trasladado = detalles_lesionados_accidentes3.cleaned_data["trasladado"]

                                nuevo_lesionado = Lesionados(
                                    id_accidente = nuevo_accidente_transito,
                                    nombres = nombre,
                                    apellidos = apellido,
                                    cedula = f"{nacionalidad}-{cedula}",
                                    edad = edad,
                                    sexo = sexo,
                                    idx = idx,
                                    descripcion = descripcion,
                                )
                                nuevo_lesionado.save()

                                if trasladado == True and traslados_accidentes3.is_valid():
                                    hospital = traslados_accidentes3.cleaned_data["hospital_trasladado"]
                                    medico = traslados_accidentes3.cleaned_data["medico_receptor"]
                                    mpps_cmt = traslados_accidentes3.cleaned_data["mpps_cmt"]

                                    nuevo_traslado_accidente = Traslado_Accidente(
                                        id_lesionado = nuevo_lesionado,
                                        hospital_trasladado = hospital,
                                        medico_receptor = medico,
                                        mpps_cmt = mpps_cmt
                                    )
                                    nuevo_traslado_accidente.save()

            if tipo_procedimiento == "9" and serv_especial.is_valid():
                descripcion = serv_especial.cleaned_data["descripcion"]
                material_utilizado = serv_especial.cleaned_data["material_utilizado"]
                status = serv_especial.cleaned_data["status"]
                tipo_servicio = serv_especial.cleaned_data["tipo_servicio"]

                tipo_servicio_instance = Tipo_servicios.objects.get(id=tipo_servicio)

                nuevo_Servicio_especial = Servicios_Especiales(
                    id_procedimientos=nuevo_procedimiento,
                    tipo_servicio = tipo_servicio_instance,
                    descripcion=descripcion,
                    material_utilizado=material_utilizado,
                    status=status
                )
                nuevo_Servicio_especial.save()

            if tipo_procedimiento == "10" and rescate_form.is_valid():
                material_utilizado = rescate_form.cleaned_data["material_utilizado"]
                status = rescate_form.cleaned_data["status"]
                id_tipo_rescate = rescate_form.cleaned_data["tipo_rescate"]


                tipo_rescate_instance = Tipo_Rescate.objects.get(id=id_tipo_rescate)

                nuevo_proc_rescate = Rescate(
                    id_procedimientos = nuevo_procedimiento,
                    material_utilizado=material_utilizado,
                    tipo_rescate = tipo_rescate_instance,
                    status=status
                )
                nuevo_proc_rescate.save()

                if id_tipo_rescate == "1" and rescate_form_animal.is_valid():
                    especie = rescate_form_animal.cleaned_data["especie"]
                    descripcion = rescate_form_animal.cleaned_data["descripcion"]

                    new_rescate_animal = Rescate_Animal(
                        id_rescate = nuevo_proc_rescate,
                        especie = especie,
                        descripcion = descripcion,
                    )
                    new_rescate_animal.save()

                    return redirect('/dashboard/')

                else:
                    rescate_form_persona.is_valid()
                    nombre_persona = rescate_form_persona.cleaned_data["nombre_persona"]
                    apellido_persona = rescate_form_persona.cleaned_data["apellido_persona"]
                    nacionalidad = rescate_form_persona.cleaned_data["nacionalidad"]
                    cedula_persona = rescate_form_persona.cleaned_data["cedula_persona"]
                    edad_persona = rescate_form_persona.cleaned_data["edad_persona"]
                    sexo_persona = rescate_form_persona.cleaned_data["sexo_persona"]
                    descripcion = rescate_form_persona.cleaned_data["descripcion"]

                    new_rescate_persona = Rescate_Persona(
                        id_rescate = nuevo_proc_rescate,
                        nombre = nombre_persona,
                        apellidos = apellido_persona,
                        cedula = f"{nacionalidad}-{cedula_persona}",
                        edad = edad_persona,
                        sexo = sexo_persona,
                        descripcion = descripcion,
                    )
                    new_rescate_persona.save()

                    return redirect('/dashboard/')

            if tipo_procedimiento == "11" and incendio_form.is_valid():
                id_tipo_incendio = incendio_form.cleaned_data["tipo_incendio"]
                descripcion = incendio_form.cleaned_data["descripcion"]
                material_utilizado = incendio_form.cleaned_data["material_utilizado"]
                status = incendio_form.cleaned_data["status"]

                tipo_incendio_instance = Tipo_Incendio.objects.get(id=id_tipo_incendio)

                nuevo_proc_incendio = Incendios(
                    id_procedimientos = nuevo_procedimiento,
                    id_tipo_incendio = tipo_incendio_instance,
                    descripcion=descripcion,
                    material_utilizado=material_utilizado,
                    status=status
                )
                nuevo_proc_incendio.save()

                check_agregar_persona = incendio_form.cleaned_data["check_agregar_persona"]

                if check_agregar_persona == True and persona_presente_form.is_valid():
                    nombre = persona_presente_form.cleaned_data["nombre"]
                    apellido = persona_presente_form.cleaned_data["apellido"]
                    nacionalidad = persona_presente_form.cleaned_data["nacionalidad"]
                    cedula = persona_presente_form.cleaned_data["cedula"]
                    edad = persona_presente_form.cleaned_data["edad"]

                    new_persona_presente = Persona_Presente(
                        id_incendio = nuevo_proc_incendio,
                        nombre = nombre,
                        apellidos = apellido,
                        cedula = f"{nacionalidad}-{cedula}",
                        edad = edad,
                    )
                    new_persona_presente.save()

                if id_tipo_incendio == "2" and detalles_vehiculo_form.is_valid():
                    modelo = detalles_vehiculo_form.cleaned_data["modelo"]
                    marca = detalles_vehiculo_form.cleaned_data["marca"]
                    color = detalles_vehiculo_form.cleaned_data["color"]
                    año = detalles_vehiculo_form.cleaned_data["año"]
                    placas = detalles_vehiculo_form.cleaned_data["placas"]

                    new_agregar_vehiculo = Detalles_Vehiculos(
                        id_vehiculo = nuevo_proc_incendio,
                        modelo = modelo,
                        marca = marca,
                        color = color,
                        año = año,
                        placas = placas,
                    )
                    new_agregar_vehiculo.save()

            if tipo_procedimiento == "12" and form_fallecido.is_valid():
                motivo_fallecimiento = form_fallecido.cleaned_data["motivo_fallecimiento"]
                nom_fallecido = form_fallecido.cleaned_data["nom_fallecido"]
                apellido_fallecido = form_fallecido.cleaned_data["apellido_fallecido"]
                nacionalidad = form_fallecido.cleaned_data["nacionalidad"]
                cedula_fallecido = form_fallecido.cleaned_data["cedula_fallecido"]
                edad = form_fallecido.cleaned_data["edad"]
                sexo = form_fallecido.cleaned_data["sexo"]
                descripcion = form_fallecido.cleaned_data["descripcion"]
                material_utilizado = form_fallecido.cleaned_data["material_utilizado"]
                status = form_fallecido.cleaned_data["status"]

                nuevo_proc_fallecido = Fallecidos(
                    id_procedimiento = nuevo_procedimiento,
                    motivo_fallecimiento = motivo_fallecimiento,
                    nombres = nom_fallecido,
                    apellidos = apellido_fallecido,
                    cedula = f"{nacionalidad}-{cedula_fallecido}",
                    edad = edad,
                    sexo = sexo,
                    descripcion=descripcion,
                    material_utilizado=material_utilizado,
                    status=status
                )
                nuevo_proc_fallecido.save()

            if tipo_procedimiento == "13" and mitigacion_riesgo_form.is_valid():
                tipo_riesgo = mitigacion_riesgo_form.cleaned_data["tipo_riesgo"]
                descripcion = mitigacion_riesgo_form.cleaned_data["descripcion"]
                material_utilizado = mitigacion_riesgo_form.cleaned_data["material_utilizado"]
                status = mitigacion_riesgo_form.cleaned_data["status"]

                tipo_riesgo_instance = Mitigacion_riesgo.objects.get(id=tipo_riesgo)

                nuevo_proc_mit = Mitigacion_Riesgos(
                    id_procedimientos = nuevo_procedimiento,
                    id_tipo_servicio = tipo_riesgo_instance,
                    descripcion=descripcion,
                    material_utilizado=material_utilizado,
                    status=status
                )
                nuevo_proc_mit.save()

            if tipo_procedimiento == "14" and evaluacion_riesgo_form.is_valid():
                tipo_riesgo = evaluacion_riesgo_form.cleaned_data["tipo_riesgo"]
                tipo_estructura = evaluacion_riesgo_form.cleaned_data["tipo_etructura"]
                descripcion = evaluacion_riesgo_form.cleaned_data["descripcion"]
                material_utilizado = evaluacion_riesgo_form.cleaned_data["material_utilizado"]
                status = evaluacion_riesgo_form.cleaned_data["status"]

                tipo_riesgo_instance = Motivo_Riesgo.objects.get(id=tipo_riesgo)

                nuevo_proc_eval = Evaluacion_Riesgo(
                    id_procedimientos = nuevo_procedimiento,
                    id_tipo_riesgo = tipo_riesgo_instance,
                    tipo_estructura = tipo_estructura,
                    descripcion=descripcion,
                    material_utilizado=material_utilizado,
                    status=status
                )

                nuevo_proc_eval.save()

                if division == "3" and tipo_procedimiento == "14" and persona_presente_eval_form.is_valid():
                    nombre = persona_presente_eval_form.cleaned_data["nombre"]
                    apellido = persona_presente_eval_form.cleaned_data["apellidos"]
                    nacionalidad = persona_presente_eval_form.cleaned_data["nacionalidad"]
                    cedula = persona_presente_eval_form.cleaned_data["cedula"]
                    telefono = persona_presente_eval_form.cleaned_data["telefono"]

                    nuevo_per_presente = Persona_Presente_Eval(
                        id_persona = nuevo_proc_eval,
                        nombre = nombre,
                        apellidos = apellido,
                        cedula = f"{nacionalidad}-{cedula}",
                        telefono = telefono,
                    )
                    nuevo_per_presente.save()

            if tipo_procedimiento == "15" and puesto_avanzada_form.is_valid():
                tipo_avanzada = puesto_avanzada_form.cleaned_data["tipo_avanzada"]
                descripcion = puesto_avanzada_form.cleaned_data["descripcion"]
                material_utilizado = puesto_avanzada_form.cleaned_data["material_utilizado"]
                status = puesto_avanzada_form.cleaned_data["status"]

                tipo_avanzada_instance = Motivo_Avanzada.objects.get(id=tipo_avanzada)

                nuevo_proc_avan = Puesto_Avanzada(
                    id_procedimientos = nuevo_procedimiento,
                    id_tipo_servicio = tipo_avanzada_instance,
                    descripcion=descripcion,
                    material_utilizado=material_utilizado,
                    status=status
                )
                nuevo_proc_avan.save()

            if tipo_procedimiento == "16" and traslados_prehospitalaria_form.is_valid():
                tipo_traslado = traslados_prehospitalaria_form.cleaned_data["tipo_traslado"]
                nombre = traslados_prehospitalaria_form.cleaned_data["nombre"]
                apellido = traslados_prehospitalaria_form.cleaned_data["apellido"]
                nacionalidad = traslados_prehospitalaria_form.cleaned_data["nacionalidad"]
                cedula = traslados_prehospitalaria_form.cleaned_data["cedula"]
                edad = traslados_prehospitalaria_form.cleaned_data["edad"]
                sexo = traslados_prehospitalaria_form.cleaned_data["sexo"]
                idx = traslados_prehospitalaria_form.cleaned_data["idx"]
                hospital_trasladado = traslados_prehospitalaria_form.cleaned_data["hospital_trasladado"]
                medico_receptor = traslados_prehospitalaria_form.cleaned_data["medico_receptor"]
                mpps_cmt = traslados_prehospitalaria_form.cleaned_data["mpps_cmt"]
                descripcion = traslados_prehospitalaria_form.cleaned_data["descripcion"]
                material_utilizado = traslados_prehospitalaria_form.cleaned_data["material_utilizado"]
                status = traslados_prehospitalaria_form.cleaned_data["status"]

                tipo_traslado_instance = Tipos_Traslado.objects.get(id=tipo_traslado)

                nuevo_proc_tras = Traslado_Prehospitalaria(
                    id_procedimiento = nuevo_procedimiento,
                    id_tipo_traslado = tipo_traslado_instance,
                    nombre = nombre,
                    apellido = apellido,
                    cedula = f"{nacionalidad}-{cedula}",
                    edad = edad,
                    sexo = sexo,
                    idx = idx,
                    hospital_trasladado = hospital_trasladado,
                    medico_receptor = medico_receptor,
                    mpps_cmt = mpps_cmt,
                    descripcion=descripcion,
                    material_utilizado=material_utilizado,
                    status=status
                )
                nuevo_proc_tras.save()

            if tipo_procedimiento == "17" and asesoramiento_form.is_valid():
                nombre_comercio = asesoramiento_form.cleaned_data["nombre_comercio"]
                rif_comercio = asesoramiento_form.cleaned_data["rif_comercio"]
                nombre = asesoramiento_form.cleaned_data["nombres"]
                apellido = asesoramiento_form.cleaned_data["apellidos"]
                nacionalidad = asesoramiento_form.cleaned_data["nacionalidad"]
                cedula = asesoramiento_form.cleaned_data["cedula"]
                sexo = asesoramiento_form.cleaned_data["sexo"]
                telefono = asesoramiento_form.cleaned_data["telefono"]
                descripcion = asesoramiento_form.cleaned_data["descripcion"]
                material_utilizado = asesoramiento_form.cleaned_data["material_utilizado"]
                status = asesoramiento_form.cleaned_data["status"]

                nuevo_proc_ase = Asesoramiento(
                    id_procedimiento = nuevo_procedimiento,
                    nombre_comercio = nombre_comercio,
                    rif_comercio = rif_comercio,
                    nombres = nombre,
                    apellidos = apellido,
                    cedula = f"{nacionalidad}-{cedula}",
                    sexo = sexo,
                    telefono = telefono,
                    descripcion=descripcion,
                    material_utilizado=material_utilizado,
                    status=status
                )
                nuevo_proc_ase.save()
                
            if tipo_procedimiento == "18" and form_inspecciones.is_valid():
                tipo_inspeccion = form_inspecciones.cleaned_data["tipo_inspeccion"]

                if tipo_inspeccion == "Prevención" and form_inspecciones_prevencion.is_valid():
                    nombre_comercio = form_inspecciones_prevencion.cleaned_data["nombre_comercio"]
                    propietario = form_inspecciones_prevencion.cleaned_data["propietario"]
                    nacionalidad = form_inspecciones_prevencion.cleaned_data["nacionalidad"]
                    cedula_propietario = form_inspecciones_prevencion.cleaned_data["cedula_propietario"]
                    descripcion = form_inspecciones_prevencion.cleaned_data["descripcion"]
                    persona_sitio_nombre = form_inspecciones_prevencion.cleaned_data["persona_sitio_nombre"]
                    persona_sitio_apellido = form_inspecciones_prevencion.cleaned_data["persona_sitio_apellido"]
                    nacionalidad2 = form_inspecciones_prevencion.cleaned_data["nacionalidad2"]
                    persona_sitio_cedula = form_inspecciones_prevencion.cleaned_data["persona_sitio_cedula"]
                    persona_sitio_telefono = form_inspecciones_prevencion.cleaned_data["persona_sitio_telefono"]
                    material_utilizado = form_inspecciones_prevencion.cleaned_data["material_utilizado"]
                    status = form_inspecciones_prevencion.cleaned_data["status"]

                    nueva_inspeccion = Inspeccion_Prevencion_Asesorias_Tecnicas (
                        id_procedimientos = nuevo_procedimiento,
                        tipo_inspeccion = tipo_inspeccion,
                        nombre_comercio = nombre_comercio,
                        propietario = propietario,
                        cedula_propietario = f"{nacionalidad}-{cedula_propietario}",
                        descripcion = descripcion,
                        persona_sitio_nombre = persona_sitio_nombre,
                        persona_sitio_apellido = persona_sitio_apellido,
                        persona_sitio_cedula = f"{nacionalidad2}-{persona_sitio_cedula}",
                        persona_sitio_telefono = persona_sitio_telefono,
                        material_utilizado = material_utilizado,
                        status = status
                    )
                
                    nueva_inspeccion.save()

                if tipo_inspeccion == "Árbol" and form_inspecciones_arbol.is_valid():
                    especie = form_inspecciones_arbol.cleaned_data["especie"]
                    altura_aprox = form_inspecciones_arbol.cleaned_data["altura_aprox"]
                    ubicacion_arbol = form_inspecciones_arbol.cleaned_data["ubicacion_arbol"]
                    persona_sitio_nombre = form_inspecciones_arbol.cleaned_data["persona_sitio_nombre"]
                    persona_sitio_apellido = form_inspecciones_arbol.cleaned_data["persona_sitio_apellido"]
                    nacionalidad = form_inspecciones_arbol.cleaned_data["nacionalidad"]
                    persona_sitio_cedula = form_inspecciones_arbol.cleaned_data["persona_sitio_cedula"]
                    persona_sitio_telefono = form_inspecciones_arbol.cleaned_data["persona_sitio_telefono"]
                    descripcion = form_inspecciones_arbol.cleaned_data["descripcion"]
                    material_utilizado = form_inspecciones_arbol.cleaned_data["material_utilizado"]
                    status = form_inspecciones_arbol.cleaned_data["status"]

                    nueva_inspeccion = Inspeccion_Arbol (
                        id_procedimientos = nuevo_procedimiento,
                        tipo_inspeccion = tipo_inspeccion,
                        especie = especie,
                        altura_aprox = altura_aprox,
                        ubicacion_arbol = ubicacion_arbol,
                        persona_sitio_nombre = persona_sitio_nombre,
                        persona_sitio_apellido = persona_sitio_apellido,
                        persona_sitio_cedula = f"{nacionalidad}-{persona_sitio_cedula}",
                        persona_sitio_telefono = persona_sitio_telefono,
                        descripcion = descripcion,
                        material_utilizado = material_utilizado,
                        status = status
                    )
                
                    nueva_inspeccion.save()

                if tipo_inspeccion == "Asesorias Tecnicas" and form_inspecciones_prevencion.is_valid():
                    nombre_comercio = form_inspecciones_prevencion.cleaned_data["nombre_comercio"]
                    propietario = form_inspecciones_prevencion.cleaned_data["propietario"]
                    nacionalidad = form_inspecciones_prevencion.cleaned_data["nacionalidad"]
                    cedula_propietario = form_inspecciones_prevencion.cleaned_data["cedula_propietario"]
                    descripcion = form_inspecciones_prevencion.cleaned_data["descripcion"]
                    persona_sitio_nombre = form_inspecciones_prevencion.cleaned_data["persona_sitio_nombre"]
                    persona_sitio_apellido = form_inspecciones_prevencion.cleaned_data["persona_sitio_apellido"]
                    nacionalidad2 = form_inspecciones_prevencion.cleaned_data["nacionalidad2"]
                    persona_sitio_cedula = form_inspecciones_prevencion.cleaned_data["persona_sitio_cedula"]
                    persona_sitio_telefono = form_inspecciones_prevencion.cleaned_data["persona_sitio_telefono"]
                    material_utilizado = form_inspecciones_prevencion.cleaned_data["material_utilizado"]
                    status = form_inspecciones_prevencion.cleaned_data["status"]

                    nueva_inspeccion = Inspeccion_Prevencion_Asesorias_Tecnicas (
                        id_procedimientos = nuevo_procedimiento,
                        tipo_inspeccion = tipo_inspeccion,
                        nombre_comercio = nombre_comercio,
                        propietario = propietario,
                        cedula_propietario = f"{nacionalidad}-{cedula_propietario}",
                        descripcion = descripcion,
                        persona_sitio_nombre = persona_sitio_nombre,
                        persona_sitio_apellido = persona_sitio_apellido,
                        persona_sitio_cedula = f"{nacionalidad2}-{persona_sitio_cedula}",
                        persona_sitio_telefono = persona_sitio_telefono,
                        material_utilizado = material_utilizado,
                        status = status
                    )
                
                    nueva_inspeccion.save()

                if tipo_inspeccion == "Habitabilidad" and form_inspecciones_habitabilidad.is_valid():
                    descripcion = form_inspecciones_habitabilidad.cleaned_data["descripcion"]
                    persona_sitio_nombre = form_inspecciones_habitabilidad.cleaned_data["persona_sitio_nombre"]
                    persona_sitio_apellido = form_inspecciones_habitabilidad.cleaned_data["persona_sitio_apellido"]
                    nacionalidad = form_inspecciones_habitabilidad.cleaned_data["nacionalidad"]
                    persona_sitio_cedula = form_inspecciones_habitabilidad.cleaned_data["persona_sitio_cedula"]
                    persona_sitio_telefono = form_inspecciones_habitabilidad.cleaned_data["persona_sitio_telefono"]
                    material_utilizado = form_inspecciones_habitabilidad.cleaned_data["material_utilizado"]
                    status = form_inspecciones_habitabilidad.cleaned_data["status"]

                    nueva_inspeccion = Inspeccion_Habitabilidad (
                        id_procedimientos = nuevo_procedimiento,
                        tipo_inspeccion = tipo_inspeccion,
                        descripcion = descripcion,
                        persona_sitio_nombre = persona_sitio_nombre,
                        persona_sitio_apellido = persona_sitio_apellido,
                        persona_sitio_cedula = f"{nacionalidad}-{persona_sitio_cedula}",
                        persona_sitio_telefono = persona_sitio_telefono,
                        material_utilizado = material_utilizado,
                        status = status
                    )
                
                    nueva_inspeccion.save()

                if tipo_inspeccion == "Otros" and form_inspecciones_otros.is_valid():
                    especifique = form_inspecciones_otros.cleaned_data["especifique"]
                    descripcion = form_inspecciones_otros.cleaned_data["descripcion"]
                    persona_sitio_nombre = form_inspecciones_otros.cleaned_data["persona_sitio_nombre"]
                    persona_sitio_apellido = form_inspecciones_otros.cleaned_data["persona_sitio_apellido"]
                    nacionalidad = form_inspecciones_otros.cleaned_data["nacionalidad"]
                    persona_sitio_cedula = form_inspecciones_otros.cleaned_data["persona_sitio_cedula"]
                    persona_sitio_telefono = form_inspecciones_otros.cleaned_data["persona_sitio_telefono"]
                    material_utilizado = form_inspecciones_otros.cleaned_data["material_utilizado"]
                    status = form_inspecciones_otros.cleaned_data["status"]

                    nueva_inspeccion = Inspeccion_Otros (
                        id_procedimientos = nuevo_procedimiento,
                        tipo_inspeccion = tipo_inspeccion,
                        especifique = especifique,
                        descripcion = descripcion,
                        persona_sitio_nombre = persona_sitio_nombre,
                        persona_sitio_apellido = persona_sitio_apellido,
                        persona_sitio_cedula = f"{nacionalidad}-{persona_sitio_cedula}",
                        persona_sitio_telefono = persona_sitio_telefono,
                        material_utilizado = material_utilizado,
                        status = status
                    )
                
                    nueva_inspeccion.save()

            if tipo_procedimiento == "19" and form_investigacion.is_valid():
                tipo_investigacion = form_investigacion.cleaned_data["tipo_investigacion"]
                tipo_siniestro = form_investigacion.cleaned_data["tipo_siniestro"]

                tipo_investigacion_instance = Tipos_Investigacion.objects.get(id=tipo_investigacion)

                new_investigacion = Investigacion(
                    id_procedimientos = nuevo_procedimiento,
                    id_tipo_investigacion = tipo_investigacion_instance,
                    tipo_siniestro = tipo_siniestro
                )
                new_investigacion.save()

                if tipo_siniestro == "Comercio" and form_inv_comercio.is_valid():
                    nombre_comercio = form_inv_comercio.cleaned_data["nombre_comercio"]
                    rif_comercio = form_inv_comercio.cleaned_data["rif_comercio"]
                    nombre_propietario = form_inv_comercio.cleaned_data["nombre_propietario"]
                    apellido_propietario = form_inv_comercio.cleaned_data["apellido_propietario"]
                    nacionalidad = form_inv_comercio.cleaned_data["nacionalidad"]
                    cedula_propietario = form_inv_comercio.cleaned_data["cedula_propietario"]
                    descripcion = form_inv_comercio.cleaned_data["descripcion"]
                    material_utilizado = form_inv_comercio.cleaned_data["material_utilizado"]
                    status = form_inv_comercio.cleaned_data["status"]

                    new_inv_comercio = Investigacion_Comercio (
                        id_investigacion = new_investigacion,
                        nombre_comercio = nombre_comercio,
                        rif_comercio = rif_comercio,
                        nombre_propietario = nombre_propietario,
                        apellido_propietario = apellido_propietario,
                        cedula_propietario = f"{nacionalidad}-{cedula_propietario}",
                        descripcion = descripcion,
                        material_utilizado = material_utilizado,
                        status = status
                    )
                    new_inv_comercio.save()

                if (tipo_siniestro == "Estructura" or tipo_siniestro == "Vivienda") and form_inv_estructura.is_valid():
                    tipo_estructura = form_inv_estructura.cleaned_data["tipo_estructura"]
                    nombre = form_inv_estructura.cleaned_data["nombre"]
                    apellido = form_inv_estructura.cleaned_data["apellido"]
                    nacionalidad = form_inv_estructura.cleaned_data["nacionalidad"]
                    cedula = form_inv_estructura.cleaned_data["cedula"]
                    descripcion = form_inv_estructura.cleaned_data["descripcion"]
                    material_utilizado = form_inv_estructura.cleaned_data["material_utilizado"]
                    status = form_inv_estructura.cleaned_data["status"]

                    new_inv_estructura = Investigacion_Estructura_Vivienda (
                        id_investigacion = new_investigacion,
                        tipo_estructura = tipo_estructura,
                        nombre = nombre,
                        apellido = apellido,
                        cedula = f"{nacionalidad}-{cedula}",
                        descripcion = descripcion,
                        material_utilizado = material_utilizado,
                        status = status
                    )
                    new_inv_estructura.save()

                if tipo_siniestro == "Vehiculo" and form_inv_vehiculo.is_valid():
                    marca = form_inv_vehiculo.cleaned_data["marca"]
                    modelo = form_inv_vehiculo.cleaned_data["modelo"]
                    color = form_inv_vehiculo.cleaned_data["color"]
                    placas = form_inv_vehiculo.cleaned_data["placas"]
                    año = form_inv_vehiculo.cleaned_data["año"]
                    nombre_propietario = form_inv_vehiculo.cleaned_data["nombre_propietario"]
                    apellido_propietario = form_inv_vehiculo.cleaned_data["apellido_propietario"]
                    nacionalidad = form_inv_vehiculo.cleaned_data["nacionalidad"]
                    cedula_propietario = form_inv_vehiculo.cleaned_data["cedula_propietario"]
                    descripcion = form_inv_vehiculo.cleaned_data["descripcion"]
                    material_utilizado = form_inv_vehiculo.cleaned_data["material_utilizado"]
                    status = form_inv_vehiculo.cleaned_data["status"]

                    new_inv_vehiculo = Investigacion_Vehiculo(
                        id_investigacion = new_investigacion,
                        marca = marca,
                        modelo = modelo,
                        color = color,
                        placas = placas,
                        año = año,
                        nombre_propietario = nombre_propietario,
                        apellido_propietario = apellido_propietario,
                        cedula_propietario = f"{nacionalidad}-{cedula_propietario}",
                        descripcion = descripcion,
                        material_utilizado = material_utilizado,
                        status = status,
                    )
                    new_inv_vehiculo.save()

            if tipo_procedimiento == "20" and reinspeccion_prevencion.is_valid():
                nombre_comercio = reinspeccion_prevencion.cleaned_data["nombre_comercio"]
                rif_comercio = reinspeccion_prevencion.cleaned_data["rif_comercio"]
                nombre = reinspeccion_prevencion.cleaned_data["nombre"]
                apellido = reinspeccion_prevencion.cleaned_data["apellidos"]
                sexo = reinspeccion_prevencion.cleaned_data["sexo"]
                nacionalidad = reinspeccion_prevencion.cleaned_data["nacionalidad"]
                cedula = reinspeccion_prevencion.cleaned_data["cedula"]
                sexo = reinspeccion_prevencion.cleaned_data["sexo"]
                telefono = reinspeccion_prevencion.cleaned_data["telefono"]
                descripcion = reinspeccion_prevencion.cleaned_data["descripcion"]
                material_utilizado = reinspeccion_prevencion.cleaned_data["material_utilizado"]
                status = reinspeccion_prevencion.cleaned_data["status"]

                nuevo_proc_reins = Reinspeccion_Prevencion(
                    id_procedimiento = nuevo_procedimiento,
                    nombre_comercio = nombre_comercio,
                    rif_comercio = rif_comercio,
                    nombre = nombre,
                    apellidos = apellido,
                    cedula = f"{nacionalidad}-{cedula}",
                    sexo = sexo,
                    telefono = telefono,
                    descripcion=descripcion,
                    material_utilizado=material_utilizado,
                    status=status
                )
                nuevo_proc_reins.save()

            if tipo_procedimiento == "21" and retencion_preventiva.is_valid():
                tipo_cilindro = retencion_preventiva.cleaned_data["tipo_cilindro"]
                capacidad = retencion_preventiva.cleaned_data["capacidad"]
                serial = retencion_preventiva.cleaned_data["serial"]
                nro_constancia_retencion = retencion_preventiva.cleaned_data["nro_constancia_retencion"]
                descripcion = retencion_preventiva.cleaned_data["descripcion"]
                material_utilizado = retencion_preventiva.cleaned_data["material_utilizado"]
                status = retencion_preventiva.cleaned_data["status"]

                nuevo_proc_reten = Retencion_Preventiva(
                    id_procedimiento = nuevo_procedimiento,
                    tipo_cilindro = tipo_cilindro,
                    capacidad = capacidad,
                    serial = serial,
                    nro_constancia_retencion = nro_constancia_retencion,
                    descripcion=descripcion,
                    material_utilizado=material_utilizado,
                    status=status
                )
                nuevo_proc_reten.save()

            if tipo_procedimiento == "22" and artificios_pirotecnico.is_valid():
                nombre_comercio = artificios_pirotecnico.cleaned_data["nombre_comercio"]
                rif_comercio = artificios_pirotecnico.cleaned_data["rif_comercio"]
                tipo_procedimiento_art = artificios_pirotecnico.cleaned_data["tipo_procedimiento"]

                tipo_procedimiento_art_instance = Tipos_Artificios.objects.get(id=tipo_procedimiento_art)

                nuevo_proc_artificio_pir = Artificios_Pirotecnicos(
                    id_procedimiento = nuevo_procedimiento,
                    nombre_comercio = nombre_comercio,
                    rif_comerciante = rif_comercio,
                    tipo_procedimiento = tipo_procedimiento_art_instance
                )

                nuevo_proc_artificio_pir.save()

                if tipo_procedimiento_art == "1" and incendio_art.is_valid():
                    id_tipo_incendio = incendio_art.cleaned_data["tipo_incendio"]
                    descripcion = incendio_art.cleaned_data["descripcion"]
                    material_utilizado = incendio_art.cleaned_data["material_utilizado"]
                    status = incendio_art.cleaned_data["status"]

                    tipo_incendio_instance = Tipo_Incendio.objects.get(id=id_tipo_incendio)

                    nuevo_proc_incendio_art = Incendios_Art(
                        id_procedimientos = nuevo_proc_artificio_pir,
                        id_tipo_incendio = tipo_incendio_instance,
                        descripcion=descripcion,
                        material_utilizado=material_utilizado,
                        status=status
                    )
                    nuevo_proc_incendio_art.save()

                    check_agregar_persona = incendio_art.cleaned_data["check_agregar_persona"]

                    if check_agregar_persona == True and persona_presente_art.is_valid():
                        nombre = persona_presente_art.cleaned_data["nombre"]
                        apellido = persona_presente_art.cleaned_data["apellido"]
                        nacionalidad = persona_presente_art.cleaned_data["nacionalidad"]
                        cedula = persona_presente_art.cleaned_data["cedula"]
                        edad = persona_presente_art.cleaned_data["edad"]

                        new_persona_presente = Persona_Presente_Art(
                            id_incendio = nuevo_proc_incendio_art,
                            nombre = nombre,
                            apellidos = apellido,
                            cedula = f"{nacionalidad}-{cedula}",
                            edad = edad,
                        )
                        new_persona_presente.save()

                    if id_tipo_incendio == "2" and detalles_vehiculo_art.is_valid():
                        modelo = detalles_vehiculo_art.cleaned_data["modelo"]
                        marca = detalles_vehiculo_art.cleaned_data["marca"]
                        color = detalles_vehiculo_art.cleaned_data["color"]
                        año = detalles_vehiculo_art.cleaned_data["año"]
                        placas = detalles_vehiculo_art.cleaned_data["placas"]

                        new_agregar_vehiculo = Detalles_Vehiculos_Art(
                            id_vehiculo = nuevo_proc_incendio_art,
                            modelo = modelo,
                            marca = marca,
                            color = color,
                            año = año,
                            placas = placas,
                        )
                        new_agregar_vehiculo.save()

                if tipo_procedimiento_art == "2" and lesionados.is_valid():
                    nombre = lesionados.cleaned_data["nombre"]
                    apellido = lesionados.cleaned_data["apellido"]
                    nacionalidad = lesionados.cleaned_data["nacionalidad"]
                    cedula = lesionados.cleaned_data["cedula"]
                    edad = lesionados.cleaned_data["edad"]
                    sexo = lesionados.cleaned_data["sexo"]
                    idx = lesionados.cleaned_data["idx"]
                    descripcion = lesionados.cleaned_data["descripcion"]
                    status = lesionados.cleaned_data["status"]


                    nuevo_lesionado_art = Lesionados_Art(
                        id_accidente = nuevo_proc_artificio_pir,
                        nombres = nombre,
                        apellidos = apellido,
                        cedula = f"{nacionalidad}-{cedula}",
                        edad = edad,
                        sexo = sexo,
                        idx = idx,
                        descripcion = descripcion,
                        status = status
                    )

                    nuevo_lesionado_art.save()

                if tipo_procedimiento_art == "3" and fallecidos_art.is_valid():
                    motivo_fallecimiento = fallecidos_art.cleaned_data["motivo_fallecimiento"]
                    nom_fallecido = fallecidos_art.cleaned_data["nom_fallecido"]
                    apellido_fallecido = fallecidos_art.cleaned_data["apellido_fallecido"]
                    nacionalidad = fallecidos_art.cleaned_data["nacionalidad"]
                    cedula_fallecido = fallecidos_art.cleaned_data["cedula_fallecido"]
                    edad = fallecidos_art.cleaned_data["edad"]
                    sexo = fallecidos_art.cleaned_data["sexo"]
                    descripcion = fallecidos_art.cleaned_data["descripcion"]
                    material_utilizado = fallecidos_art.cleaned_data["material_utilizado"]
                    status = fallecidos_art.cleaned_data["status"]

                    nuevo_proc_fallecido_art = Fallecidos_Art(
                        id_procedimiento = nuevo_proc_artificio_pir,
                        motivo_fallecimiento = motivo_fallecimiento,
                        nombres = nom_fallecido,
                        apellidos = apellido_fallecido,
                        cedula = f"{nacionalidad}-{cedula_fallecido}",
                        edad = edad,
                        sexo = sexo,
                        descripcion=descripcion,
                        material_utilizado=material_utilizado,
                        status=status
                    )
                    nuevo_proc_fallecido_art.save()

            if tipo_procedimiento == "23" and inspeccion_artificios_pir.is_valid():
                nombre_comercio = inspeccion_artificios_pir.cleaned_data["nombre_comercio"]
                rif_comercio = inspeccion_artificios_pir.cleaned_data["rif_comercio"]
                nombre_encargado = inspeccion_artificios_pir.cleaned_data["nombre_encargado"]
                apellido_encargado = inspeccion_artificios_pir.cleaned_data["apellido_encargado"]
                nacionalidad = inspeccion_artificios_pir.cleaned_data["nacionalidad"]
                cedula_encargado = inspeccion_artificios_pir.cleaned_data["cedula_encargado"]
                sexo = inspeccion_artificios_pir.cleaned_data["sexo"]
                descripcion = inspeccion_artificios_pir.cleaned_data["descripcion"]
                material_utilizado = inspeccion_artificios_pir.cleaned_data["material_utilizado"]
                status = inspeccion_artificios_pir.cleaned_data["status"]

                nueva_inspeccion_art = Inspeccion_Establecimiento_Art(
                    id_proc_artificio = nuevo_procedimiento,
                    nombre_comercio = nombre_comercio,
                    rif_comercio = rif_comercio,
                    encargado_nombre = nombre_encargado,
                    encargado_apellidos = apellido_encargado,
                    encargado_cedula = f"{nacionalidad}-{cedula_encargado}",
                    encargado_sexo = sexo,
                    descripcion = descripcion,
                    material_utilizado = material_utilizado,
                    status = status
                )

                nueva_inspeccion_art.save()

            if tipo_procedimiento == "24" and form_valoracion_medica.is_valid():
                nombre = form_valoracion_medica.cleaned_data["nombre"]
                apellido = form_valoracion_medica.cleaned_data["apellido"]
                nacionalidad = form_valoracion_medica.cleaned_data["nacionalidad"]
                cedula = form_valoracion_medica.cleaned_data["cedula"]
                edad = form_valoracion_medica.cleaned_data["edad"]
                sexo = form_valoracion_medica.cleaned_data["sexo"]
                telefono = form_valoracion_medica.cleaned_data["telefono"]
                descripcion = form_valoracion_medica.cleaned_data["descripcion"]
                material_utilizado = form_valoracion_medica.cleaned_data["material_utilizado"]
                status = form_valoracion_medica.cleaned_data["status"]

                new_valoracion_medica = Valoracion_Medica(
                    id_procedimientos = nuevo_procedimiento,
                    nombre = nombre,
                    apellido = apellido,
                    cedula = f"{nacionalidad}-{cedula}",
                    edad = edad,
                    sexo = sexo,
                    telefono = telefono,
                    descripcion = descripcion,
                    material_utilizado = material_utilizado,
                    status = status
                )
                new_valoracion_medica.save()

            if tipo_procedimiento == "25" and form_jornada_medica.is_valid():
                nombre_jornada = form_jornada_medica.cleaned_data["nombre_jornada"]
                cantidad_personas_atendidas = form_jornada_medica.cleaned_data["cant_personas_aten"]
                descripcion = form_jornada_medica.cleaned_data["descripcion"]
                material_utilizado = form_jornada_medica.cleaned_data["material_utilizado"]
                status = form_jornada_medica.cleaned_data["status"]

                new_jornada_medica = Jornada_Medica(
                    id_procedimientos = nuevo_procedimiento,
                    nombre_jornada = nombre_jornada,
                    cant_personas_aten = cantidad_personas_atendidas,
                    descripcion = descripcion,
                    material_utilizado = material_utilizado,
                    status = status
                )
                new_jornada_medica.save()

            if (tipo_procedimiento == "26" or tipo_procedimiento == "27" or tipo_procedimiento == "28" or tipo_procedimiento == "29" or tipo_procedimiento == "30" or tipo_procedimiento == "31" or tipo_procedimiento == "32" or tipo_procedimiento == "33" or tipo_procedimiento == "34") and form_detalles_enfermeria.is_valid():
                nombre = form_detalles_enfermeria.cleaned_data["nombre"]
                apellido = form_detalles_enfermeria.cleaned_data["apellido"]
                nacionalidad = form_detalles_enfermeria.cleaned_data["nacionalidad"]
                cedula = form_detalles_enfermeria.cleaned_data["cedula"]
                edad = form_detalles_enfermeria.cleaned_data["edad"]
                sexo = form_detalles_enfermeria.cleaned_data["sexo"]
                telefono = form_detalles_enfermeria.cleaned_data["telefono"]
                descripcion = form_detalles_enfermeria.cleaned_data["descripcion"]
                material_utilizado = form_detalles_enfermeria.cleaned_data["material_utilizado"]
                status = form_detalles_enfermeria.cleaned_data["status"]

                new_detalles_enfermeria = Detalles_Enfermeria(
                    id_procedimientos = nuevo_procedimiento,
                    nombre = nombre,
                    apellido = apellido,
                    cedula = f"{nacionalidad}-{cedula}",
                    edad = edad,
                    sexo = sexo,
                    telefono = telefono,
                    descripcion = descripcion,
                    material_utilizado = material_utilizado,
                    status = status
                )
                new_detalles_enfermeria.save()

            if (tipo_procedimiento == "35" or tipo_procedimiento == "36" or tipo_procedimiento == "37" or tipo_procedimiento == "38" or tipo_procedimiento == "39" or tipo_procedimiento == "40" or tipo_procedimiento == "41") and form_detalles_psicologia.is_valid():
                nombre = form_detalles_psicologia.cleaned_data["nombre"]
                apellido = form_detalles_psicologia.cleaned_data["apellido"]
                nacionalidad = form_detalles_psicologia.cleaned_data["nacionalidad"]
                cedula = form_detalles_psicologia.cleaned_data["cedula"]
                edad = form_detalles_psicologia.cleaned_data["edad"]
                sexo = form_detalles_psicologia.cleaned_data["sexo"]
                descripcion = form_detalles_psicologia.cleaned_data["descripcion"]
                material_utilizado = form_detalles_psicologia.cleaned_data["material_utilizado"]
                status = form_detalles_psicologia.cleaned_data["status"]

                new_detalles_psicologia = Procedimientos_Psicologia(
                    id_procedimientos = nuevo_procedimiento,
                    nombre = nombre,
                    apellido = apellido,
                    cedula = f"{nacionalidad}-{cedula}",
                    edad = edad,
                    sexo = sexo,
                    descripcion = descripcion,
                    material_utilizado = material_utilizado,
                    status = status
                )
                new_detalles_psicologia.save()

            # Redirige a /dashboard/ después de guardar los datos
            return redirect('/dashboard/')
    else:
        form = SelectorDivision(prefix='form1')
        form2 = SeleccionarInfo(prefix='form2')
        form3 = Datos_Ubicacion(prefix='form3')
        form4 = Selecc_Tipo_Procedimiento(prefix='form4')
        abast_agua = formulario_abastecimiento_agua(prefix='abast_agua')
        apoyo_unid = Formulario_apoyo_unidades(prefix='apoyo_unid')
        guard_prev = Formulario_guardia_prevencion(prefix='guard_prev')
        atend_no_efec = Formulario_atendido_no_efectuado(prefix='atend_no_efec')
        desp_seguridad = Formulario_despliegue_seguridad(prefix='desp_seguridad')
        fals_alarm = Formulario_falsa_alarma(prefix='fals_alarm')
        serv_especial = Formulario_Servicios_Especiales(prefix='serv_especial')
        form_fallecido = Formulario_Fallecidos(prefix='form_fallecido')
        rescate_form = Formulario_Rescate(prefix='rescate_form')
        incendio_form = Formulario_Incendio(prefix='incendio_form')
        atenciones_paramedicas = Formulario_Atenciones_Paramedicas(prefix='atenciones_paramedicas')

        emergencias_medicas = Formulario_Emergencias_Medicas(prefix='emergencias_medicas')
        traslados_emergencias = Formulario_Traslados(prefix='traslados_emergencias')

        persona_presente_form = Formulario_Persona_Presente(prefix='persona_presente_form')
        detalles_vehiculo_form = Formulario_Detalles_Vehiculos_Incendio(prefix='detalles_vehiculo_form')

        formulario_accidentes_transito = Formulario_Accidentes_Transito(prefix='formulario_accidentes_transito')
        detalles_vehiculo_accidentes = Formulario_Detalles_Vehiculos(prefix='detalles_vehiculos_accidentes')
        detalles_lesionados_accidentes = Formulario_Detalles_Lesionados(prefix='detalles_lesionados_accidentes')
        detalles_lesionados_accidentes2 = Formulario_Detalles_Lesionados2(prefix='detalles_lesionados_accidentes2')
        detalles_lesionados_accidentes3 = Formulario_Detalles_Lesionados3(prefix='detalles_lesionados_accidentes3')
        traslados_accidentes = Formulario_Traslado_Accidente(prefix='traslados_accidentes')
        traslados_accidentes2 = Formulario_Traslado_Accidente2(prefix='traslados_accidentes2')
        traslados_accidentes3 = Formulario_Traslado_Accidente3(prefix='traslados_accidentes3')
        detalles_vehiculo_accidentes2 = Formulario_Detalles_Vehiculos2(prefix='detalles_vehiculos_accidentes2')
        detalles_vehiculo_accidentes3 = Formulario_Detalles_Vehiculos3(prefix='detalles_vehiculos_accidentes3')

        rescate_form_persona = Formulario_Rescate_Persona(prefix='rescate_form_persona')
        rescate_form_animal = Formulario_Rescate_Animal(prefix='rescate_form_animal')

        evaluacion_riesgo_form = Forulario_Evaluacion_Riesgo(prefix='evaluacion_riesgo_form')
        mitigacion_riesgo_form = Formulario_Mitigacion_Riesgos(prefix='mitigacion_riesgo_form')

        puesto_avanzada_form = Formulario_Puesto_Avanzada(prefix='puesto_avanzada_form')
        traslados_prehospitalaria_form = Formulario_Traslados_Prehospitalaria(prefix='traslados_prehospitalaria_form')
        asesoramiento_form = Formulario_Asesoramiento(prefix='asesoramiento_form')
        persona_presente_eval_form = Formularia_Persona_Presente_Eval(prefix='persona_presente_eval_form')
        reinspeccion_prevencion = Formulario_Reinspeccion_Prevencion(prefix='reinspeccion_prevencion')
        retencion_preventiva = Formulario_Retencion_Preventiva(prefix='retencion_preventiva')

        artificios_pirotecnico = Formulario_Artificios_Pirotecnicos(prefix='artificios_pirotecnico')
        lesionados = Formulario_Lesionado(prefix='lesionados')
        incendio_art = Formulario_Incendio_Art(prefix='incendio_art')
        persona_presente_art = Formulario_Persona_Presente_Art(prefix='persona_presente_art')
        detalles_vehiculo_art = Formulario_Detalles_Vehiculos_Incendio_Art(prefix='detalles_vehiculo_art')
        fallecidos_art = Formulario_Fallecidos_Art(prefix='fallecidos_art')
        inspeccion_artificios_pir = Formulario_Inspeccion_Establecimiento_Art(prefix='inspeccion_artificios_pir')
        form_enfermeria = Formulario_Enfermeria(prefix='form_enfermeria')
        servicios_medicos = Formulario_Servicios_medicos(prefix='form_servicios_medicos')
        psicologia = Formulario_psicologia(prefix='form_psicologia')
        capacitacion = Formulario_capacitacion(prefix='form_capacitacion')
        form_valoracion_medica = Formulario_Valoracion_Medica(prefix='form_valoracion_medica')
        form_jornada_medica = Formulario_Jornada_Medica(prefix='form_jornada_medica')
        form_detalles_enfermeria = Formulario_Detalles_Enfermeria(prefix='form_detalles_enfermeria')
        form_detalles_psicologia = Formulario_Procedimientos_Psicologia(prefix='form_detalles_psicologia')

        form_capacitacion = Formulario_Capacitacion_Proc(prefix='form_capacitacion')
        form_frente_preventivo = Formulario_Frente_Preventivo(prefix='form_frente_preventivo')

        form_inspecciones = Formulario_Inspecciones(prefix='form_inspecciones')
        form_inspecciones_prevencion = Formulario_Inspeccion_Prevencion_Asesorias_Tecnicas(prefix='form_inspecciones_prevencion')
        form_inspecciones_habitabilidad = Formulario_Inspeccion_Habitabilidad(prefix='form_inspecciones_habitabilidad')
        form_inspecciones_arbol = Formulario_Inspeccion_Arbol(prefix='form_inspecciones_arbol')
        form_inspecciones_otros = Formulario_Inspeccion_Otros(prefix='form_inspecciones_otros')

        form_investigacion = Formulario_Investigacion(prefix='form_investigacion')
        form_inv_vehiculo = Formulario_Investigacion_Vehiculo(prefix='form_inv_vehiculo')
        form_inv_comercio = Formulario_Investigacion_Comercio(prefix='form_inv_comercio')
        form_inv_estructura = Formulario_Investigacion_Estructura_Vivienda(prefix='form_inv_estructura')

    return render(request, "procedimientos.html", {
        "user": user,
        "jerarquia": user["jerarquia"],
        "nombres": user["nombres"],
        "apellidos": user["apellidos"],
        "form": form,
        "form2": form2,
        "form3": form3,
        "form4": form4,
        "errors": result,
        "form_abastecimiento_agua": abast_agua,
        "form_apoyo_unidades": apoyo_unid,
        "form_guardia_prevencion": guard_prev,
        "form_atendido_no_efectuado": atend_no_efec,
        "form_despliegue_seguridad": desp_seguridad,
        "form_falsa_alarma": fals_alarm,
        "form_servicios_especiales": serv_especial,
        "form_fallecido": form_fallecido,
        "rescate_form": rescate_form,
        "rescate_form_animal": rescate_form_animal,
        "rescate_form_persona": rescate_form_persona,
        "incendio_form": incendio_form,
        "persona_presente_form": persona_presente_form,
        "detalles_vehiculo_form": detalles_vehiculo_form,
        "atenciones_paramedicas": atenciones_paramedicas,
        "emergencias_medicas": emergencias_medicas,
        "traslados_emergencias": traslados_emergencias,
        "formulario_accidentes_transito": formulario_accidentes_transito,
        "detalles_vehiculo_accidentes":  detalles_vehiculo_accidentes,
        "detalles_vehiculo_accidentes2":  detalles_vehiculo_accidentes2,
        "detalles_vehiculo_accidentes3":  detalles_vehiculo_accidentes3,
        "detalles_lesionados_accidentes": detalles_lesionados_accidentes,
        "detalles_lesionados_accidentes2": detalles_lesionados_accidentes2,
        "detalles_lesionados_accidentes3": detalles_lesionados_accidentes3,
        "traslados_accidentes": traslados_accidentes,
        "traslados_accidentes2": traslados_accidentes2,
        "traslados_accidentes3": traslados_accidentes3,
        "evaluacion_riesgo_form": evaluacion_riesgo_form,
        "mitigacion_riesgo_form": mitigacion_riesgo_form,
        "puesto_avanzada_form": puesto_avanzada_form,
        "traslados_prehospitalaria_form": traslados_prehospitalaria_form,
        "asesoramiento_form": asesoramiento_form,
        "persona_presente_eval_form": persona_presente_eval_form,
        "reinspeccion_prevencion": reinspeccion_prevencion,
        "retencion_preventiva": retencion_preventiva,
        "artificios_pirotecnico": artificios_pirotecnico,
        "lesionados": lesionados,
        "incendio_art": incendio_art,
        "persona_presente_art": persona_presente_art,
        "detalles_vehiculo_art": detalles_vehiculo_art,
        "fallecidos_art": fallecidos_art,
        "inspeccion_artificios_pir": inspeccion_artificios_pir,
        "form_enfermeria": form_enfermeria,
        "servicios_medicos" : servicios_medicos,
        "psicologia" : psicologia,
        "capacitacion" : capacitacion,
        "valoracion_medica": form_valoracion_medica,
        "form_detalles_enfermeria": form_detalles_enfermeria,
        "form_detalles_psicologia": form_detalles_psicologia,
        "form_capacitacion": form_capacitacion,
        "form_frente_preventivo": form_frente_preventivo,
        "jornada_medica": form_jornada_medica,
        "form_inspecciones": form_inspecciones,
        "form_inspecciones_prevencion": form_inspecciones_prevencion,
        "form_inspecciones_habitabilidad": form_inspecciones_habitabilidad,
        "form_inspecciones_arbol": form_inspecciones_arbol,
        "form_inspecciones_otros": form_inspecciones_otros,
        "form_investigacion": form_investigacion,
        "form_inv_vehiculo": form_inv_vehiculo,
        "form_inv_comercio": form_inv_comercio,
        "form_inv_estructura": form_inv_estructura,
        })

# Vista de la seccion de Estadisticas
def View_Estadisticas(request):
    user = request.session.get('user')
    if not user:
            return redirect('/')

    return render(request, "estadisticas.html", {
        "user": user,
        "jerarquia": user["jerarquia"],
        "nombres": user["nombres"],
        "apellidos": user["apellidos"],
    })
# Vista de la Seccion de Operaciones

def View_Operaciones(request):
    user = request.session.get('user')
    if not user:
            return redirect('/')

    datos = Procedimientos.objects.filter(id_division=2)
    total = datos.count()

    # Obtener la fecha de hoy
    hoy = datetime.now().date()

    # Filtrar procedimientos con la fecha de hoy
    fechas = datos.values_list("fecha", flat=True)
    procedimientos_hoy = [fecha for fecha in fechas if fecha == hoy]

    hoy = len(procedimientos_hoy)

    if request.method == 'POST':
        data = json.loads(request.body)
        id = data.get('id')
        procedimiento = get_object_or_404(Procedimientos, id=id)
        try:
            procedimiento.delete()
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})

    datos = list(datos)[::-1]

    return render(request, "Divisiones/operaciones.html", {
        "user": user,
        "jerarquia": user["jerarquia"],
        "nombres": user["nombres"],
        "apellidos": user["apellidos"],
        "datos": datos,
        "total": total,
        "hoy": hoy
    })
# Vista de la Seccion de Operaciones

def View_Rescate(request):
    user = request.session.get('user')
    if not user:
            return redirect('/')

    datos = Procedimientos.objects.filter(id_division = 1)

    total = datos.count()

    # Obtener la fecha de hoy
    hoy = datetime.now().date()

    # Filtrar procedimientos con la fecha de hoy
    fechas = datos.values_list("fecha", flat=True)
    procedimientos_hoy = [fecha for fecha in fechas if fecha == hoy]

    hoy = len(procedimientos_hoy)

    if request.method == 'POST':
        data = json.loads(request.body)
        id = data.get('id')
        procedimiento = get_object_or_404(Procedimientos, id=id)
        try:
            procedimiento.delete()
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})

    datos = list(datos)[::-1]

    return render(request, "Divisiones/rescate.html", {
        "user": user,
        "jerarquia": user["jerarquia"],
        "nombres": user["nombres"],
        "apellidos": user["apellidos"],
        "datos": datos,
        "total": total,
        "hoy": hoy
    })

def View_Prevencion(request):
    user = request.session.get('user')
    if not user:
            return redirect('/')

    datos = Procedimientos.objects.filter(id_division = 3)
    total = datos.count()

    # Obtener la fecha de hoy
    hoy = datetime.now().date()

    # Filtrar procedimientos con la fecha de hoy
    fechas = datos.values_list("fecha", flat=True)
    procedimientos_hoy = [fecha for fecha in fechas if fecha == hoy]

    hoy = len(procedimientos_hoy)

    if request.method == 'POST':
        data = json.loads(request.body)
        id = data.get('id')
        procedimiento = get_object_or_404(Procedimientos, id=id)
        try:
            procedimiento.delete()
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
        
    datos = list(datos)[::-1]

    return render(request, "Divisiones/prevencion.html", {
        "user": user,
        "jerarquia": user["jerarquia"],
        "nombres": user["nombres"],
        "apellidos": user["apellidos"],
        "datos": datos,
        "total": total,
        "hoy": hoy
    })

def View_grumae(request):
    user = request.session.get('user')
    if not user:
            return redirect('/')

    datos = Procedimientos.objects.filter(id_division = 4)

    total = datos.count()

    # Obtener la fecha de hoy
    hoy = datetime.now().date()

    # Filtrar procedimientos con la fecha de hoy
    fechas = datos.values_list("fecha", flat=True)
    procedimientos_hoy = [fecha for fecha in fechas if fecha == hoy]

    hoy = len(procedimientos_hoy)

    if request.method == 'POST':
        data = json.loads(request.body)
        id = data.get('id')
        procedimiento = get_object_or_404(Procedimientos, id=id)
        try:
            procedimiento.delete()
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
        
    datos = list(datos)[::-1]

    return render(request, "Divisiones/grumae.html", {
        "user": user,
        "jerarquia": user["jerarquia"],
        "nombres": user["nombres"],
        "apellidos": user["apellidos"],
        "datos": datos,
        "total": total,
        "hoy": hoy
    })

def View_prehospitalaria(request):
    user = request.session.get('user')
    if not user:
            return redirect('/')

    datos = Procedimientos.objects.filter(id_division = 5)

    total = datos.count()

    # Obtener la fecha de hoy
    hoy = datetime.now().date()

    # Filtrar procedimientos con la fecha de hoy
    fechas = datos.values_list("fecha", flat=True)
    procedimientos_hoy = [fecha for fecha in fechas if fecha == hoy]

    hoy = len(procedimientos_hoy)


    if request.method == 'POST':
        data = json.loads(request.body)
        id = data.get('id')
        procedimiento = get_object_or_404(Procedimientos, id=id)
        try:
            procedimiento.delete()
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
        
    datos = list(datos)[::-1]

    return render(request, "Divisiones/prehospitalaria.html", {
        "user": user,
        "jerarquia": user["jerarquia"],
        "nombres": user["nombres"],
        "apellidos": user["apellidos"],
        "datos": datos,
        "total": total,
        "hoy": hoy
    })

def View_capacitacion(request):
    user = request.session.get('user')
    if not user:
            return redirect('/')

    datos = Procedimientos.objects.filter(id_division = 9)

    total = datos.count()

    # Obtener la fecha de hoy
    hoy = datetime.now().date()

    # Filtrar procedimientos con la fecha de hoy
    fechas = datos.values_list("fecha", flat=True)
    procedimientos_hoy = [fecha for fecha in fechas if fecha == hoy]

    hoy = len(procedimientos_hoy)

    if request.method == 'POST':
        data = json.loads(request.body)
        id = data.get('id')
        procedimiento = get_object_or_404(Procedimientos, id=id)
        try:
            procedimiento.delete()
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
        
    datos = list(datos)[::-1]

    return render(request, "Divisiones/capacitacion.html", {
        "user": user,
        "jerarquia": user["jerarquia"],
        "nombres": user["nombres"],
        "apellidos": user["apellidos"],
        "datos": datos,
        "total": total,
        "hoy": hoy
    })

def View_enfermeria(request):
    user = request.session.get('user')
    if not user:
            return redirect('/')

    datos = Procedimientos.objects.filter(id_division = 6)

    total = datos.count()

    # Obtener la fecha de hoy
    hoy = datetime.now().date()

    # Filtrar procedimientos con la fecha de hoy
    fechas = datos.values_list("fecha", flat=True)
    procedimientos_hoy = [fecha for fecha in fechas if fecha == hoy]

    hoy = len(procedimientos_hoy)

    if request.method == 'POST':
        data = json.loads(request.body)
        id = data.get('id')
        procedimiento = get_object_or_404(Procedimientos, id=id)
        try:
            procedimiento.delete()
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
        
    datos = list(datos)[::-1]

    return render(request, "Divisiones/enfermeria.html", {
        "user": user,
        "jerarquia": user["jerarquia"],
        "nombres": user["nombres"],
        "apellidos": user["apellidos"],
        "datos": datos,
        "total": total,
        "hoy": hoy
    })

def View_serviciosmedicos(request):
    user = request.session.get('user')
    if not user:
            return redirect('/')

    datos = Procedimientos.objects.filter(id_division = 7)

    total = datos.count()

    # Obtener la fecha de hoy
    hoy = datetime.now().date()

    # Filtrar procedimientos con la fecha de hoy
    fechas = datos.values_list("fecha", flat=True)
    procedimientos_hoy = [fecha for fecha in fechas if fecha == hoy]

    hoy = len(procedimientos_hoy)


    if request.method == 'POST':
        data = json.loads(request.body)
        id = data.get('id')
        procedimiento = get_object_or_404(Procedimientos, id=id)
        try:
            procedimiento.delete()
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
        
    datos = list(datos)[::-1]

    return render(request, "Divisiones/serviciosmedicos.html", {
        "user": user,
        "jerarquia": user["jerarquia"],
        "nombres": user["nombres"],
        "apellidos": user["apellidos"],
        "datos": datos,
        "total": total,
        "hoy": hoy
    })

def View_psicologia(request):
    user = request.session.get('user')
    if not user:
            return redirect('/')

    datos = Procedimientos.objects.filter(id_division = 8)

    total = datos.count()

    # Obtener la fecha de hoy
    hoy = datetime.now().date()

    # Filtrar procedimientos con la fecha de hoy
    fechas = datos.values_list("fecha", flat=True)
    procedimientos_hoy = [fecha for fecha in fechas if fecha == hoy]

    hoy = len(procedimientos_hoy)


    if request.method == 'POST':
        data = json.loads(request.body)
        id = data.get('id')
        procedimiento = get_object_or_404(Procedimientos, id=id)
        try:
            procedimiento.delete()
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
        
    datos = list(datos)[::-1]

    return render(request, "Divisiones/psicologia.html", {
        "user": user,
        "jerarquia": user["jerarquia"],
        "nombres": user["nombres"],
        "apellidos": user["apellidos"],
        "datos": datos,
        "total": total,
        "hoy": hoy
    })

def tabla_general(request):
    user = request.session.get('user')
    if not user:
        return redirect('/')

    # Filtra los procedimientos de las divisiones 1 a 5
    divisiones = range(1, 6)
    datos_combined = Procedimientos.objects.filter(id_division__in=divisiones).order_by('-fecha')  # Orden descendente

    # Corrige el conteo
    total = datos_combined.count()

    # Obtener la fecha de hoy
    hoy_inicio = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    hoy_fin = hoy_inicio + timedelta(days=1)


    # Filtrar procedimientos con la fecha de hoy
    procedimientos_hoy = datos_combined.filter(fecha__gte=hoy_inicio, fecha__lt=hoy_fin)
    hoy_count = procedimientos_hoy.count()

    if request.method == 'POST':
        data = json.loads(request.body)
        id = data.get('id')
        procedimiento = get_object_or_404(Procedimientos, id=id)
        try:
            procedimiento.delete()
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})

    return render(request, "tablageneral.html", {
        "user": user,
        "jerarquia": user["jerarquia"],
        "nombres": user["nombres"],
        "apellidos": user["apellidos"],
        "datos": datos_combined,  # Ya en orden descendente
        "total": total,
        "hoy": hoy_count
    })

# Funcion para eliminar (NO TOCAR)
# def eliminar_procedimiento(request):
#     if request.method == 'POST':
#         data = json.loads(request.body)
#         id = data.get('id')
#         procedimiento = get_object_or_404(Procedimientos, id=id)
#         try:
#             procedimiento.delete()
#             return JsonResponse({'success': True})
#         except Exception as e:
#             return JsonResponse({'success': False, 'error': str(e)})

def obtener_procedimiento(request, id):
    procedimiento = get_object_or_404(Procedimientos, pk=id)

    division = procedimiento.id_division.division

    if division == "Rescate" or division == "Operaciones" or division == "Prevencion" or division == "GRUMAE" or division == "PreHospitalaria":
        data = {
            'id': procedimiento.id,
            'division': procedimiento.id_division.division,
            'solicitante': f"{procedimiento.id_solicitante.jerarquia} {procedimiento.id_solicitante.nombres} {procedimiento.id_solicitante.apellidos}",
            'solicitante_externo': procedimiento.solicitante_externo,
            'jefe_comision': f"{procedimiento.id_jefe_comision.jerarquia} {procedimiento.id_jefe_comision.nombres} {procedimiento.id_jefe_comision.apellidos}",
            'unidad': procedimiento.unidad.nombre_unidad,
            'efectivos': procedimiento.efectivos_enviados,
            'parroquia': procedimiento.id_parroquia.parroquia,
            'municipio': procedimiento.id_municipio.municipio,
            'direccion': procedimiento.direccion,
            'fecha': procedimiento.fecha,
            'hora': procedimiento.hora,
            'tipo_procedimiento': procedimiento.id_tipo_procedimiento.tipo_procedimiento,
        }
    
    if division == "Enfermeria":
        data = {
            'id': procedimiento.id,
            'division': procedimiento.id_division.division,
            'dependencia': procedimiento.dependencia,
            'solicitante_externo': procedimiento.solicitante_externo,
            'parroquia': procedimiento.id_parroquia.parroquia,
            'municipio': procedimiento.id_municipio.municipio,
            'direccion': procedimiento.direccion,
            'fecha': procedimiento.fecha,
            'hora': procedimiento.hora,
            'tipo_procedimiento': procedimiento.id_tipo_procedimiento.tipo_procedimiento,
        }

    if division == "Servicios Medicos":
        data = {
            'id': procedimiento.id,
            'division': procedimiento.id_division.division,
            'tipo_servicio': procedimiento.tipo_servicio,
            'solicitante_externo': procedimiento.solicitante_externo,
            'efectivos': procedimiento.efectivos_enviados,
            'parroquia': procedimiento.id_parroquia.parroquia,
            'municipio': procedimiento.id_municipio.municipio,
            'direccion': procedimiento.direccion,
            'fecha': procedimiento.fecha,
            'hora': procedimiento.hora,
            'tipo_procedimiento': procedimiento.id_tipo_procedimiento.tipo_procedimiento,
        }

    if division == "Psicologia":
            data = {
                'id': procedimiento.id,
                'division': procedimiento.id_division.division,
                'solicitante_externo': procedimiento.solicitante_externo,
                'parroquia': procedimiento.id_parroquia.parroquia,
                'municipio': procedimiento.id_municipio.municipio,
                'direccion': procedimiento.direccion,
                'fecha': procedimiento.fecha,
                'hora': procedimiento.hora,
                'tipo_procedimiento': procedimiento.id_tipo_procedimiento.tipo_procedimiento,
            }
    
    if division == "Capacitacion":
        data = {
            'id': procedimiento.id,
            'division': procedimiento.id_division.division,
            'solicitante': f"{procedimiento.id_solicitante.jerarquia} {procedimiento.id_solicitante.nombres} {procedimiento.id_solicitante.apellidos}",
            'solicitante_externo': procedimiento.solicitante_externo,
            'jefe_comision': f"{procedimiento.id_jefe_comision.jerarquia} {procedimiento.id_jefe_comision.nombres} {procedimiento.id_jefe_comision.apellidos}",
            'dependencia': procedimiento.dependencia,
            'parroquia': procedimiento.id_parroquia.parroquia,
            'municipio': procedimiento.id_municipio.municipio,
            'direccion': procedimiento.direccion,
            'fecha': procedimiento.fecha,
            'hora': procedimiento.hora,
            'tipo_procedimiento': procedimiento.id_tipo_procedimiento.tipo_procedimiento,
        }
    
    if str(procedimiento.id_tipo_procedimiento.id) == "1":
        detalle_procedimiento = get_object_or_404(Abastecimiento_agua, id_procedimiento=id)

        data = dict(data,
                    ente_suministrado = detalle_procedimiento.id_tipo_servicio.nombre_institucion,
                    nombres = detalle_procedimiento.nombres,
                    apellidos = detalle_procedimiento.apellidos,
                    cedula = detalle_procedimiento.cedula,
                    ltrs_agua = detalle_procedimiento.ltrs_agua,
                    personas_atendidas = detalle_procedimiento.personas_atendidas,
                    descripcion = detalle_procedimiento.descripcion,
                    material_utilizado = detalle_procedimiento.material_utilizado,
                    status = detalle_procedimiento.status)

    if str(procedimiento.id_tipo_procedimiento.id) == "2":
        detalle_procedimiento = get_object_or_404(Apoyo_Unidades, id_procedimiento=id)
        data = dict(data,
                    tipo_apoyo = detalle_procedimiento.id_tipo_apoyo.tipo_apoyo,
                    unidad_apoyada = detalle_procedimiento.unidad_apoyada,
                    descripcion = detalle_procedimiento.descripcion,
                    material_utilizado = detalle_procedimiento.material_utilizado,
                    status = detalle_procedimiento.status,
                    )

    if str(procedimiento.id_tipo_procedimiento.id) == "3":
        detalle_procedimiento = get_object_or_404(Guardia_prevencion, id_procedimiento=id)
        data = dict(data,
                    motivo_prevencion = detalle_procedimiento.id_motivo_prevencion.motivo,
                    descripcion = detalle_procedimiento.descripcion,
                    material_utilizado = detalle_procedimiento.material_utilizado,
                    status = detalle_procedimiento.status,
                    )

    if str(procedimiento.id_tipo_procedimiento.id) == "4":
        detalle_procedimiento = get_object_or_404(Atendido_no_Efectuado, id_procedimiento=id)
        data = dict(data,
                    descripcion = detalle_procedimiento.descripcion,
                    material_utilizado = detalle_procedimiento.material_utilizado,
                    status = detalle_procedimiento.status,
                    )

    if str(procedimiento.id_tipo_procedimiento.id) == "5":
        detalle_procedimiento = get_object_or_404(Despliegue_Seguridad, id_procedimiento=id)
        data = dict(data,
                    motivo_despliegue = detalle_procedimiento. motivo_despliegue.motivo,
                    descripcion = detalle_procedimiento.descripcion,
                    material_utilizado = detalle_procedimiento.material_utilizado,
                    status = detalle_procedimiento.status,
                    )

    if str(procedimiento.id_tipo_procedimiento.id) == "6":
        detalle_procedimiento = get_object_or_404(Falsa_Alarma, id_procedimiento=id)
        data = dict(data,
                    motivo_alarma = detalle_procedimiento.motivo_alarma.motivo,
                    descripcion = detalle_procedimiento.descripcion,
                    material_utilizado = detalle_procedimiento.material_utilizado,
                    status = detalle_procedimiento.status,
                    )

    if str(procedimiento.id_tipo_procedimiento.id) == "7":
        # Obtener el detalle del procedimiento
        detalle_procedimiento = get_object_or_404(Atenciones_Paramedicas, id_procedimientos=id)

        # Agregar detalles del procedimiento a los datos
        data = dict(data,
                    tipo_atencion=detalle_procedimiento.tipo_atencion,
                    )

        if detalle_procedimiento.tipo_atencion == "Emergencias Medicas":
            emergencia = Emergencias_Medicas.objects.get(id_atencion=detalle_procedimiento.id)
            data = dict(data,
                emergencia = True,
                nombres = emergencia.nombres,
                apellidos = emergencia.apellidos,
                cedula = emergencia.cedula,
                edad = emergencia.edad,
                sexo = emergencia.sexo,
                idx = emergencia.idx,
                descripcion = emergencia.descripcion,
                material_utilizado = emergencia.material_utilizado,
                status = emergencia.status,
            )

            if Traslado.objects.filter(id_lesionado=emergencia.id).exists():
                traslado = Traslado.objects.get(id_lesionado = emergencia.id)

                data = dict(data,
                            traslado = True,
                            hospital = traslado.hospital_trasladado,
                            medico = traslado.medico_receptor,
                            mpps_cmt = traslado.mpps_cmt,
                        )

        if detalle_procedimiento.tipo_atencion == "Accidentes de Transito":
            accidente = Accidentes_Transito.objects.get(id_atencion=detalle_procedimiento.id)
            data = dict(data,
                accidente = True,
                tipo_accidente=accidente.tipo_de_accidente.tipo_accidente,
                cantidad_lesionados=accidente.cantidad_lesionados,
                material_utilizado=accidente.material_utilizado,
                status=accidente.status,
            )

            # Filtrar todos los vehículos relacionados con el accidente
            vehiculos = Detalles_Vehiculos_Accidente.objects.filter(id_vehiculo=accidente.id)

            # Si hay vehículos, recopilarlos en una lista
            if vehiculos:
                data = dict(data,
                    vehiculo = True
                )
                vehiculos_list = []
                for vehiculo in vehiculos:
                    vehiculos_list.append({
                        'marca': vehiculo.marca,
                        'modelo': vehiculo.modelo,
                        'color': vehiculo.color,
                        'año': vehiculo.año,
                        'placas': vehiculo.placas,
                        # Añade aquí otros campos que necesites
                    })
                data['vehiculos'] = vehiculos_list  # Agrega la lista de vehículos a 'data'
            else:
                data['vehiculos'] = []  # O puedes omitir esta línea si prefieres no agregar la clave

            # Filtrar los lesionados asociados al accidente
            lesionados = Lesionados.objects.filter(id_accidente=accidente.id)

            # Si hay lesionados, recopilarlos en una lista
            if lesionados:
                data = dict(data,
                    lesionado = True
                )
                lesionados_list = []
                for lesionado in lesionados:
                    lesionado_data = {
                        'nombre': lesionado.nombres,
                        'apellidos': lesionado.apellidos,
                        'cedula': lesionado.cedula,
                        'edad': lesionado.edad,
                        'sexo': lesionado.sexo,
                        'idx': lesionado.idx,
                        'descripcion': lesionado.descripcion,
                        # Añade aquí otros campos que necesites
                    }

                    # Filtrar traslados asociados a cada lesionado
                    traslados = Traslado_Accidente.objects.filter(id_lesionado=lesionado.id)

                    # Si hay traslados, añadirlos a los datos del lesionado
                    if traslados:
                        traslados_list = []
                        for traslado in traslados:
                            traslados_list.append({
                                'hospital': traslado.hospital_trasladado,
                                'medico': traslado.medico_receptor,
                                'mpps_cmt': traslado.mpps_cmt,
                            })
                        lesionado_data['traslados'] = traslados_list
                    else:
                        lesionado_data['traslados'] = []

                    # Añadir cada lesionado a la lista
                    lesionados_list.append(lesionado_data)

                data['lesionados'] = lesionados_list  # Agregar la lista de lesionados a 'data'
            else:
                data['lesionados'] = []  # Si no hay lesionados, agregar una lista vacía

    if str(procedimiento.id_tipo_procedimiento.id) == "9":
        detalle_procedimiento = get_object_or_404(Servicios_Especiales, id_procedimientos=id)
        data = dict(data,
                    tipo_servicio = detalle_procedimiento.tipo_servicio.serv_especiales,
                    descripcion = detalle_procedimiento.descripcion,
                    material_utilizado = detalle_procedimiento.material_utilizado,
                    status = detalle_procedimiento.status,
                    )

    if str(procedimiento.id_tipo_procedimiento.id) == "10":
        detalle_procedimiento = get_object_or_404(Rescate, id_procedimientos=id)
        data = dict(data,
                    material_utilizado = detalle_procedimiento.material_utilizado,
                    status = detalle_procedimiento.status,
                    tipo_rescate = detalle_procedimiento.tipo_rescate.tipo_rescate,
                    )

        if detalle_procedimiento.tipo_rescate.tipo_rescate == "Rescate de Animal":
            detalle_tipo_rescate = get_object_or_404(Rescate_Animal, id_rescate=detalle_procedimiento.id)
            data = dict(data,
                        especie = detalle_tipo_rescate.especie,
                        descripcion = detalle_tipo_rescate.descripcion,
                        )

        else:
            detalle_tipo_rescate = get_object_or_404(Rescate_Persona, id_rescate=detalle_procedimiento.id)
            data = dict(data,
                        nombres = detalle_tipo_rescate.nombre,
                        apellidos = detalle_tipo_rescate.apellidos,
                        cedula = detalle_tipo_rescate.cedula,
                        edad = detalle_tipo_rescate.edad,
                        sexo = detalle_tipo_rescate.sexo,
                        descripcion = detalle_tipo_rescate.descripcion,
                        )

    if str(procedimiento.id_tipo_procedimiento.id) == "11":
      # Obtener el detalle del procedimiento
      detalle_procedimiento = get_object_or_404(Incendios, id_procedimientos=id)

      # Agregar detalles del procedimiento a los datos
      data = dict(data,
                  tipo_incendio=detalle_procedimiento.id_tipo_incendio.tipo_incendio,
                  descripcion=detalle_procedimiento.descripcion,
                  status=detalle_procedimiento.status,
                  material_utilizado=detalle_procedimiento.material_utilizado,
                )

      if Persona_Presente.objects.filter(id_incendio=detalle_procedimiento.id).exists():
          persona_presente_detalles = Persona_Presente.objects.get(id_incendio=detalle_procedimiento.id)
          data.update({
              "persona": True,
              "nombre": persona_presente_detalles.nombre,
              "apellidos": persona_presente_detalles.apellidos,
              "cedula": persona_presente_detalles.cedula,
              "edad": persona_presente_detalles.edad,
          })
      else:
          pass

      if Detalles_Vehiculos.objects.filter(id_vehiculo=detalle_procedimiento.id).exists():
          vehiculo_detalles = Detalles_Vehiculos.objects.get(id_vehiculo=detalle_procedimiento.id)
          data.update({
              "vehiculo": True,
              "modelo": vehiculo_detalles.modelo,
              "marca": vehiculo_detalles.marca,
              "color": vehiculo_detalles.color,
              "año": vehiculo_detalles.año,
              "placas": vehiculo_detalles.placas,
          })
      else:
          pass

    if str(procedimiento.id_tipo_procedimiento.id) == "12":
        detalle_procedimiento = get_object_or_404(Fallecidos, id_procedimiento=id)
        data = dict(data,
                    motivo_fallecimiento = detalle_procedimiento.motivo_fallecimiento,
                    nombres = detalle_procedimiento.nombres,
                    apellidos = detalle_procedimiento.apellidos,
                    cedula = detalle_procedimiento.cedula,
                    edad = detalle_procedimiento.edad,
                    sexo = detalle_procedimiento.sexo,
                    descripcion = detalle_procedimiento.descripcion,
                    material_utilizado = detalle_procedimiento.material_utilizado,
                    status = detalle_procedimiento.status,
                    )

    if str(procedimiento.id_tipo_procedimiento.id) == "13":
        detalle_procedimiento = get_object_or_404(Mitigacion_Riesgos, id_procedimientos=id)
        data = dict(data,
                    tipo_servicio = detalle_procedimiento.id_tipo_servicio.tipo_servicio,
                    descripcion = detalle_procedimiento.descripcion,
                    material_utilizado = detalle_procedimiento.material_utilizado,
                    status = detalle_procedimiento.status,
                    )

    if str(procedimiento.id_tipo_procedimiento.id) == "14":
        detalle_procedimiento = get_object_or_404(Evaluacion_Riesgo, id_procedimientos=id)
        data = dict(data,
                    tipo_de_evaluacion = detalle_procedimiento.id_tipo_riesgo.tipo_riesgo,
                    descripcion = detalle_procedimiento.descripcion,
                    material_utilizado = detalle_procedimiento.material_utilizado,
                    status = detalle_procedimiento.status,
                    )

        if str(detalle_procedimiento.id_procedimientos.id_division) == "Prevencion":
            detalle_persona = get_object_or_404(Persona_Presente_Eval, id_persona=detalle_procedimiento.id)
            data = dict(data,
                        nombre = detalle_persona.nombre,
                        apellido = detalle_persona.apellidos,
                        cedula = detalle_persona.cedula,
                        telefono = detalle_persona.telefono,
                        )
        if detalle_procedimiento.tipo_estructura:
            data = dict(data,
                        tipo_estructura = detalle_procedimiento.tipo_estructura)

        else:
            data = dict(data,
                        tipo_estructura = "")

    if str(procedimiento.id_tipo_procedimiento.id) == "15":
        detalle_procedimiento = get_object_or_404(Puesto_Avanzada, id_procedimientos=id)
        data = dict(data,
                    tipo_de_servicio = detalle_procedimiento.id_tipo_servicio.tipo_servicio,
                    descripcion = detalle_procedimiento.descripcion,
                    material_utilizado = detalle_procedimiento.material_utilizado,
                    status = detalle_procedimiento.status,
                    )

    if str(procedimiento.id_tipo_procedimiento.id) == "16":
        detalle_procedimiento = get_object_or_404(Traslado_Prehospitalaria, id_procedimiento=id)
        data = dict(data,
                    traslado = detalle_procedimiento.id_tipo_traslado.tipo_traslado,
                    descripcion = detalle_procedimiento.descripcion,
                    material_utilizado = detalle_procedimiento.material_utilizado,
                    status = detalle_procedimiento.status,
                    nombre = detalle_procedimiento.nombre,
                    apellido = detalle_procedimiento.apellido,
                    cedula = detalle_procedimiento.cedula,
                    edad = detalle_procedimiento.edad,
                    sexo = detalle_procedimiento.sexo,
                    idx = detalle_procedimiento.idx,
                    hospital = detalle_procedimiento.hospital_trasladado,
                    medico = detalle_procedimiento.medico_receptor,
                    mpps = detalle_procedimiento.mpps_cmt
                    )

    if str(procedimiento.id_tipo_procedimiento.id) == "17":
        detalle_procedimiento = get_object_or_404(Asesoramiento, id_procedimiento=id)
        data = dict(data,
                    nombre_comercio = detalle_procedimiento.nombre_comercio,
                    rif_comercio = detalle_procedimiento.rif_comercio,
                    nombre = detalle_procedimiento.nombres,
                    apellido = detalle_procedimiento.apellidos,
                    cedula = detalle_procedimiento.cedula,
                    sexo = detalle_procedimiento.sexo,
                    telefono = detalle_procedimiento.telefono,
                    descripcion = detalle_procedimiento.descripcion,
                    material_utilizado = detalle_procedimiento.material_utilizado,
                    status = detalle_procedimiento.status,
                    )

    if str(procedimiento.id_tipo_procedimiento.id) == "18":  # Supongamos que 18 es el ID de Procedimiento
        # Diccionario para mapear tipos de inspección a sus modelos
        inspection_models = {
            "Prevención": Inspeccion_Prevencion_Asesorias_Tecnicas,
            "Habitabilidad": Inspeccion_Habitabilidad,
            "Otros": Inspeccion_Otros,
            "Arbol": Inspeccion_Arbol
        }

        # Intentar obtener la instancia de inspección
        for tipo_inspeccion, model_class in inspection_models.items():
            try:
                detalle_procedimiento = model_class.objects.get(id_procedimientos=id)

                # Actualizar datos en función del tipo de inspección
                data.update({
                    "tipo_inspeccion": detalle_procedimiento.tipo_inspeccion,
                    "persona_sitio_nombre": detalle_procedimiento.persona_sitio_nombre,
                    "persona_sitio_apellido": detalle_procedimiento.persona_sitio_apellido,
                    "persona_sitio_cedula": detalle_procedimiento.persona_sitio_cedula,
                    "persona_sitio_telefono": detalle_procedimiento.persona_sitio_telefono,
                    "material_utilizado": detalle_procedimiento.material_utilizado,
                    "status": detalle_procedimiento.status,
                })

                # Actualizar campos específicos según el tipo de inspección
                if tipo_inspeccion == "Prevención":
                    data.update({
                        "nombre_comercio": detalle_procedimiento.nombre_comercio,
                        "propietario": detalle_procedimiento.propietario,
                        "cedula_propietario": detalle_procedimiento.cedula_propietario,
                        "descripcion": detalle_procedimiento.descripcion,
                    })
                elif tipo_inspeccion == "Habitabilidad":
                    data.update({
                        "descripcion": detalle_procedimiento.descripcion,
                    })
                elif tipo_inspeccion == "Otros":
                    data.update({
                        "especifique": detalle_procedimiento.especifique,
                        "descripcion": detalle_procedimiento.descripcion,
                    })
                elif tipo_inspeccion == "Arbol":
                    data.update({
                        "especie": detalle_procedimiento.especie,
                        "altura_aprox": detalle_procedimiento.altura_aprox,
                        "ubicacion_arbol": detalle_procedimiento.ubicacion_arbol,
                        "descripcion": detalle_procedimiento.descripcion,
                    })

                # Salir del ciclo una vez que se haya encontrado y procesado la inspección
                break

            except model_class.DoesNotExist:
                # Si no se encuentra la inspección, continuar con el siguiente tipo
                continue

    if str(procedimiento.id_tipo_procedimiento.id) == "19":
        investigacion = get_object_or_404(Investigacion, id_procedimientos=id)
        data.update({
            "tipo_investigacion": investigacion.id_tipo_investigacion.tipo_investigacion,
            "tipo_siniestro": investigacion.tipo_siniestro,
        })
        
        if investigacion.tipo_siniestro == "Vehiculo":
            vehiculo = Investigacion_Vehiculo.objects.filter(id_investigacion=investigacion).first()
            if vehiculo:
                data.update({
                    "marca": vehiculo.marca,
                    "modelo": vehiculo.modelo,
                    "color": vehiculo.color,
                    "placas": vehiculo.placas,
                    "año": vehiculo.año,
                    "nombre_propietario": vehiculo.nombre_propietario,
                    "apellido_propietario": vehiculo.apellido_propietario,
                    "cedula_propietario": vehiculo.cedula_propietario,
                    "descripcion": vehiculo.descripcion,
                    "material_utilizado": vehiculo.material_utilizado,
                    "status": vehiculo.status,
                })

        elif investigacion.tipo_siniestro == "Comercio":
            comercio = Investigacion_Comercio.objects.filter(id_investigacion=investigacion).first()
            if comercio:
                data.update({
                    "nombre_comercio_investigacion": comercio.nombre_comercio,
                    "rif_comercio_investigacion": comercio.rif_comercio,
                    "nombre_propietario_comercio": comercio.nombre_propietario,
                    "apellido_propietario_comercio": comercio.apellido_propietario,
                    "cedula_propietario_comercio": comercio.cedula_propietario,
                    "descripcion_comercio": comercio.descripcion,
                    "material_utilizado_comercio": comercio.material_utilizado,
                    "status_comercio": comercio.status,
                })

        elif investigacion.tipo_siniestro == "Estructura" or investigacion.tipo_siniestro == "Vivienda":
            estructura = Investigacion_Estructura_Vivienda.objects.filter(id_investigacion=investigacion).first()
            if estructura:
                data.update({
                    "tipo_estructura": estructura.tipo_estructura,
                    "nombre_propietario_estructura": estructura.nombre,
                    "apellido_propietario_estructura": estructura.apellido,
                    "cedula_propietario_estructura": estructura.cedula,
                    "descripcion_estructura": estructura.descripcion,
                    "material_utilizado_estructura": estructura.material_utilizado,
                    "status_estructura": estructura.status,
                })

    if str(procedimiento.id_tipo_procedimiento.id) == "20":
        detalle_procedimiento = get_object_or_404(Reinspeccion_Prevencion, id_procedimiento=id)
        data = dict(data,
                    nombre_comercio = detalle_procedimiento.nombre_comercio,
                    rif_comercio = detalle_procedimiento.rif_comercio,
                    nombre = detalle_procedimiento.nombre,
                    apellido = detalle_procedimiento.apellidos,
                    cedula = detalle_procedimiento.cedula,
                    sexo = detalle_procedimiento.sexo,
                    telefono = detalle_procedimiento.telefono,
                    descripcion = detalle_procedimiento.descripcion,
                    material_utilizado = detalle_procedimiento.material_utilizado,
                    status = detalle_procedimiento.status,
                    )

    if str(procedimiento.id_tipo_procedimiento.id) == "21":
        detalle_procedimiento = get_object_or_404(Retencion_Preventiva, id_procedimiento=id)
        data = dict(data,
                    tipo_retencion = "Cilindro GLP",
                    tipo_cilindro = detalle_procedimiento.tipo_cilindro,
                    capacidad = detalle_procedimiento.capacidad,
                    serial = detalle_procedimiento.serial,
                    nro_constancia = detalle_procedimiento.nro_constancia_retencion,
                    descripcion = detalle_procedimiento.descripcion,
                    material_utilizado = detalle_procedimiento.material_utilizado,
                    status = detalle_procedimiento.status,
                    )

    if str(procedimiento.id_tipo_procedimiento.id) == "22":
        detalle_procedimiento = get_object_or_404(Artificios_Pirotecnicos, id_procedimiento=id)

        data = dict(data,
                   nombre_comercio = detalle_procedimiento.nombre_comercio,
                    rif_comercio = detalle_procedimiento.rif_comerciante,
                    tipo_procedimiento_art = detalle_procedimiento.tipo_procedimiento.tipo,
                )

        if detalle_procedimiento.tipo_procedimiento.id == 1:
            incendio = get_object_or_404(Incendios_Art, id_procedimientos=detalle_procedimiento.id)
            data.update({
                'tipo_incendio': incendio.id_tipo_incendio.tipo_incendio,
                'descripcion': incendio.descripcion,
                'material_utilizado': incendio.material_utilizado,
                'status': incendio.status,
            })

            try:
                if get_object_or_404(Persona_Presente_Art, id_incendio=incendio.id):
                    persona = get_object_or_404(Persona_Presente_Art, id_incendio=incendio.id)
                    data = dict(data,
                                person = True,
                                nombre = persona.nombre,
                                apellidos = persona.apellidos,
                                cedula = persona.cedula,
                                edad = persona.edad,
                                )
            except: 
                pass

            if incendio.id_tipo_incendio.tipo_incendio == "Incendio de Vehiculo":
                vehiculo = get_object_or_404(Detalles_Vehiculos_Art, id_vehiculo=incendio.id)
                data = dict(data,
                            carro = True,
                            modelo = vehiculo.modelo,
                            marca = vehiculo.marca,
                            color = vehiculo.color,
                            año = vehiculo.año,
                            placas = vehiculo.placas,
                            )

        if detalle_procedimiento.tipo_procedimiento.id == 2:
            lesionado = get_object_or_404(Lesionados_Art, id_accidente=detalle_procedimiento.id)
            data.update({
                'nombres': lesionado.nombres,
                'apellidos': lesionado.apellidos,
                'cedula': lesionado.cedula,
                'edad': lesionado.edad,
                'sexo': lesionado.sexo,
                'idx': lesionado.idx,
                'descripcion': lesionado.descripcion,
                'status': lesionado.status,
            })

        if detalle_procedimiento.tipo_procedimiento.id == 3:
            fallecido = get_object_or_404(Fallecidos_Art, id_procedimiento=detalle_procedimiento.id)
            data.update({
                'motivo_fallecimiento': fallecido.motivo_fallecimiento,
                'nombres': fallecido.nombres,
                'apellidos': fallecido.apellidos,
                'cedula': fallecido.cedula,
                'edad': fallecido.edad,
                'sexo': fallecido.sexo,
                'descripcion': fallecido.descripcion,
                'material_utilizado': fallecido.material_utilizado,
                'status': fallecido.status,
            })

    if str(procedimiento.id_tipo_procedimiento.id) == "23":
        detalle_procedimiento = get_object_or_404(Inspeccion_Establecimiento_Art, id_proc_artificio=id)
        data = dict(data,
                    nombre_comercio = detalle_procedimiento.nombre_comercio,
                    rif_comercio = detalle_procedimiento.rif_comercio,
                    encargado_nombre = detalle_procedimiento.encargado_nombre,
                    encargado_apellidos = detalle_procedimiento.encargado_apellidos,
                    encargado_cedula = detalle_procedimiento.encargado_cedula,
                    encargado_sexo = detalle_procedimiento.encargado_sexo,
                    descripcion = detalle_procedimiento.descripcion,
                    material_utilizado = detalle_procedimiento.material_utilizado,
                    status = detalle_procedimiento.status,
                    )

    if str(procedimiento.id_tipo_procedimiento.id) == "24":
        detalles = get_object_or_404(Valoracion_Medica, id_procedimientos = id)

        data = dict(data,
                    nombres = detalles.nombre,
                    apellidos = detalles.apellido,
                    cedula = detalles.cedula,
                    edad = detalles.edad,
                    sexo = detalles.sexo,
                    telefono = detalles.telefono,
                    descripcion = detalles.descripcion,
                    material_utilizado = detalles.material_utilizado,
                    status = detalles.status,
                    )
        
    if str(procedimiento.id_tipo_procedimiento.id) == "25":
        detalles = get_object_or_404(Jornada_Medica, id_procedimientos = id)

        data = dict(data,
                    nombre_jornada = detalles.nombre_jornada,
                    cant_personas = detalles.cant_personas_aten,
                    descripcion = detalles.descripcion,
                    material_utilizado = detalles.material_utilizado,
                    status = detalles.status,
                    )

    if str(procedimiento.id_tipo_procedimiento.id) == "26" or str(procedimiento.id_tipo_procedimiento.id) == "27" or str(procedimiento.id_tipo_procedimiento.id) == "28" or str(procedimiento.id_tipo_procedimiento.id) == "29" or str(procedimiento.id_tipo_procedimiento.id) == "30" or str(procedimiento.id_tipo_procedimiento.id) == "31" or str(procedimiento.id_tipo_procedimiento.id) == "32" or str(procedimiento.id_tipo_procedimiento.id) == "33" or str(procedimiento.id_tipo_procedimiento.id) == "34":
        detalles = get_object_or_404(Detalles_Enfermeria, id_procedimientos = id)

        data = dict(data,
                    nombres = detalles.nombre,
                    apellidos = detalles.apellido,
                    cedula = detalles.cedula,
                    edad = detalles.edad,
                    sexo = detalles.sexo,
                    telefono = detalles.telefono,
                    descripcion = detalles.descripcion,
                    material_utilizado = detalles.material_utilizado,
                    status = detalles.status,
                    )
        
    if str(procedimiento.id_tipo_procedimiento.id) == "35" or str(procedimiento.id_tipo_procedimiento.id) == "36" or str(procedimiento.id_tipo_procedimiento.id) == "37" or str(procedimiento.id_tipo_procedimiento.id) == "38" or str(procedimiento.id_tipo_procedimiento.id) == "39" or str(procedimiento.id_tipo_procedimiento.id) == "40" or str(procedimiento.id_tipo_procedimiento.id) == "41":
        detalles = get_object_or_404(Procedimientos_Psicologia, id_procedimientos = id)

        data = dict(data,
                    nombres = detalles.nombre,
                    apellidos = detalles.apellido,
                    cedula = detalles.cedula,
                    edad = detalles.edad,
                    sexo = detalles.sexo,
                    descripcion = detalles.descripcion,
                    material_utilizado = detalles.material_utilizado,
                    status = detalles.status,
                    )
        
    if str(procedimiento.id_tipo_procedimiento.id) == "45":
        if procedimiento.dependencia == "Capacitacion":
            detalles = get_object_or_404(Procedimientos_Capacitacion, id_procedimientos = id)

            data = dict(data,
                    tipo_capacitacion = detalles.tipo_capacitacion,
                    tipo_clasificacion = detalles.tipo_clasificacion,
                    personas_beneficiadas = detalles.personas_beneficiadas,
                    descripcion = detalles.descripcion,
                    material_utilizado = detalles.material_utilizado,
                    status = detalles.status
                    )
        
        if procedimiento.dependencia == "Frente Preventivo":
            detalles = get_object_or_404(Procedimientos_Frente_Preventivo, id_procedimientos = id)

            data = dict(data,
                    nombre_actividad = detalles.nombre_actividad,
                    estrategia = detalles.estrategia,
                    personas_beneficiadas = detalles.personas_beneficiadas,
                    descripcion = detalles.descripcion,
                    material_utilizado = detalles.material_utilizado,
                    status = detalles.status
                    )
        
    return JsonResponse(data)